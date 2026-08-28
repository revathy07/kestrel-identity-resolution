#!/usr/bin/env python3
"""Verify generated identity-resolution data against Assessment No. 6's data brief.

The verifier is intentionally read-only. It validates observable output and marks
conditions that need generator provenance as WARN rather than pretending they passed.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import math
import re
import sys
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any, Iterable

try:
    from openpyxl import load_workbook
except ImportError as exc:  # pragma: no cover - environment error
    raise SystemExit("Missing dependency: pip install openpyxl") from exc


EXPECTED_SCHEMAS = {
    "app_users": {
        "account_id", "email", "phone", "first_name", "last_name", "dob",
        "device_id", "device_type", "signup_ts", "country",
    },
    "store_customers": {
        "customer_id", "customer_email_address", "contact_no", "first", "last",
        "dob", "device", "device_type", "line1", "line2", "city", "postcode",
        "country",
    },
    "ticketing": {
        "booking_id", "full_name", "email", "phone", "guest", "device_id",
        "event_ts", "created_ts",
    },
    "subscriptions": {
        "subscription_id", "email", "subscriber_name", "billing_name", "payment_token", "start_date",
        "country",
    },
    "social_logins": {"provider", "provider_id"},
}

FILES = {
    "app_users": "app_users.csv",
    "store_customers": "store_customers.csv",
    "ticketing": "ticketing.jl",
    "subscriptions": "subscriptions.xlsx",
    "social_logins": "social_logins.json",
}

ID_FIELDS = {
    "app_users": "account_id",
    "store_customers": "customer_id",
    "ticketing": "booking_id",
    "subscriptions": "subscription_id",
    "social_logins": "provider_id",
}


@dataclass
class Check:
    status: str
    name: str
    detail: str


@dataclass
class SourceStats:
    count: int = 0
    fields: set[str] = field(default_factory=set)
    missing: Counter[str] = field(default_factory=Counter)
    exact_extra: int = 0
    near_extra: int = 0
    malformed: int = 0
    bot_rows: int = 0
    qa_rows: int = 0
    impossible_rows: int = 0
    late_rows: int = 0
    late_eligible: int = 0
    poison: Counter[str] = field(default_factory=Counter)
    text_features: Counter[str] = field(default_factory=Counter)
    timestamp_formats: dict[str, set[str]] = field(default_factory=lambda: defaultdict(set))
    categories: dict[str, set[str]] = field(default_factory=lambda: defaultdict(set))
    provider_shapes: dict[str, set[tuple[str, ...]]] = field(default_factory=lambda: defaultdict(set))
    join_refs: dict[str, list[str]] = field(default_factory=lambda: defaultdict(list))
    special: Counter[str] = field(default_factory=Counter)
    ids: set[str] = field(default_factory=set)
    numeric_ids: set[int] = field(default_factory=set)
    order_decreased: bool = False
    _seen_hashes: set[bytes] = field(default_factory=set)
    _id_hashes: dict[str, set[bytes]] = field(default_factory=lambda: defaultdict(set))
    _previous_numeric_id: int | None = None


class Audit:
    def __init__(self) -> None:
        self.checks: list[Check] = []

    def add(self, passed: bool, name: str, detail: str) -> None:
        self.checks.append(Check("PASS" if passed else "FAIL", name, detail))

    def warn(self, name: str, detail: str) -> None:
        self.checks.append(Check("WARN", name, detail))

    def print(self) -> int:
        for item in self.checks:
            print(f"[{item.status:4}] {item.name}: {item.detail}")
        counts = Counter(item.status for item in self.checks)
        print(
            f"\nSummary: {counts['PASS']} passed, {counts['FAIL']} failed, "
            f"{counts['WARN']} warnings"
        )
        return 1 if counts["FAIL"] else 0


def fingerprint(record: dict[str, Any]) -> bytes:
    encoded = json.dumps(record, sort_keys=True, ensure_ascii=False, default=str).encode("utf-8")
    return hashlib.blake2b(encoded, digest_size=16).digest()


def is_missing_identifier(value: Any) -> bool:
    if value is None:
        return True
    text = str(value).strip()
    while len(text) >= 2 and text[0] == text[-1] and text[0] in {'"', "'"}:
        text = text[1:-1].strip()
    return text.casefold() in {"", "null", "none"}


def normalize_email(value: Any) -> str:
    if is_missing_identifier(value):
        return ""
    text = str(value).splitlines()[0].split(' "quoted')[0].strip()
    if is_missing_identifier(text):
        return ""
    text = text.lower()
    if "@" in text:
        local, domain = text.split("@", 1)
        if domain == "brand-example.test":
            local = local.split("+", 1)[0].replace(".", "")
        text = f"{local}@{domain}"
    return text


def naive_tokens(record: dict[str, Any]) -> set[str]:
    """Reconstruct generator-independent evidence used by the documented naive matcher."""
    result: set[str] = set()
    for key in ("email", "customer_email_address", "verified_email"):
        value = normalize_email(record.get(key))
        if value:
            result.add("email:" + value)
    if record.get("hashed_email"):
        result.add("hash:" + str(record["hashed_email"]))
    for key in ("phone", "contact_no"):
        if record.get(key):
            digits = re.sub(r"\D", "", str(record[key]))
            if digits in {"0000000000", "9999999999"}:
                result.add("phone:" + digits)
            elif digits:
                result.add("phone:" + digits.removeprefix("999").lstrip("0"))
    for key in ("device_id", "device"):
        if record.get(key):
            result.add("device:" + str(record[key]))
    if record.get("account_id"):
        result.add("account:" + str(int(str(record["account_id"]))))
    if record.get("app_account_ref"):
        result.add("account:" + str(int(str(record["app_account_ref"]))))
    if record.get("payment_token"):
        result.add("payment:" + str(record["payment_token"]))
    if record.get("dob") in {"1900-01-01", "1970-01-01", "01-01-1900", "01-01-1970"}:
        result.add("dob:" + str(record["dob"]))
    city = record.get("city")
    if city:
        if record.get("first_name") or record.get("last_name"):
            name = f"{record.get('first_name', '')} {record.get('last_name', '')}"
        elif record.get("first") or record.get("last"):
            name = f"{record.get('first', '')} {record.get('last', '')}"
        else:
            name = str(record.get("full_name") or record.get("display_name") or "")
        first_line = name.splitlines()
        name = (first_line[0] if first_line else "").split(' "quoted')[0].split(" hello ")[0].replace("😊", "")
        clean_name = re.sub(r"[^a-z]", "", name.lower())
        clean_city = re.sub(r"[^a-z]", "", str(city).lower())
        if clean_name and clean_city:
            result.add(f"name_city:{clean_name}:{clean_city}")
    return result


def record_has_poison(record: dict[str, Any]) -> bool:
    values = [value for value in record.values() if isinstance(value, (str, int, float, bool, type(None)))]
    return (
        any(value in {"0000000000", "9999999999", "1900-01-01", "1970-01-01", "01-01-1900", "01-01-1970", "bookings@events.example", "KIOSK-DEVICE-1"} for value in values)
        or any(isinstance(value, str) and value.lower().endswith("@staff.test") for value in values)
    )


class NaiveClusterTracker:
    def __init__(self) -> None:
        self.parent: list[int] = []
        self.size: list[int] = []
        self.owner: dict[str, int] = {}
        self.poisoned: list[int] = []

    def find(self, node: int) -> int:
        while self.parent[node] != node:
            self.parent[node] = self.parent[self.parent[node]]
            node = self.parent[node]
        return node

    def union(self, left: int, right: int) -> None:
        left_root, right_root = self.find(left), self.find(right)
        if left_root == right_root:
            return
        if self.size[left_root] < self.size[right_root]:
            left_root, right_root = right_root, left_root
        self.parent[right_root] = left_root
        self.size[left_root] += self.size[right_root]

    def add(self, record: dict[str, Any]) -> None:
        index = len(self.parent)
        self.parent.append(index)
        self.size.append(1)
        if record_has_poison(record):
            self.poisoned.append(index)
        for evidence in naive_tokens(record):
            previous = self.owner.setdefault(evidence, index)
            if previous != index:
                self.union(index, previous)

    def largest_poisoned_cluster(self) -> int:
        return max((self.size[self.find(index)] for index in self.poisoned), default=0)


def timestamp_format(value: str) -> str | None:
    value = value.strip()
    patterns = (
        (r"^\d{2}-\d{2}-\d{4} \d{2}:\d{2}:\d{2}$", "DD-MM-YYYY"),
        (r"^\d{4}/\d{2}/\d{2} \d{2}:\d{2}:\d{2}$", "YYYY/MM/DD"),
        (r"^\d{2}-\d{2}-\d{2} \d{2}:\d{2}$", "MM-DD-YY"),
        (r"^\d{4}-\d{2}-\d{2}T.*(?:Z|[+-]\d{2}:\d{2})$", "ISO_OFFSET"),
        (r"^\d{13}$", "EPOCH_MS"),
        (r"^\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}$", "LOCAL_TEXT"),
    )
    return next((label for pattern, label in patterns if re.match(pattern, value)), None)


def parse_timestamp(value: Any) -> datetime | None:
    if not isinstance(value, str) or not value.strip():
        return None
    value = value.strip()
    fmt = timestamp_format(value)
    try:
        if fmt == "DD-MM-YYYY":
            parsed = datetime.strptime(value, "%d-%m-%Y %H:%M:%S")
        elif fmt == "YYYY/MM/DD":
            parsed = datetime.strptime(value, "%Y/%m/%d %H:%M:%S")
        elif fmt == "MM-DD-YY":
            parsed = datetime.strptime(value, "%m-%d-%y %H:%M")
        elif fmt == "EPOCH_MS":
            return datetime.fromtimestamp(int(value) / 1000, tz=timezone.utc)
        elif fmt == "LOCAL_TEXT":
            parsed = datetime.strptime(value, "%Y-%m-%d %H:%M:%S")
        elif fmt == "ISO_OFFSET":
            parsed = datetime.fromisoformat(value.replace("Z", "+00:00"))
        else:
            return None
        return parsed.astimezone(timezone.utc) if parsed.tzinfo else parsed.replace(tzinfo=timezone.utc)
    except (ValueError, OverflowError):
        return None


def iter_csv(path: Path) -> Iterable[dict[str, Any]]:
    with path.open("r", encoding="utf-8", newline="") as handle:
        yield from csv.DictReader(handle)


def iter_jsonl(path: Path) -> Iterable[dict[str, Any]]:
    with path.open("r", encoding="utf-8") as handle:
        for line_number, line in enumerate(handle, 1):
            if line.strip():
                try:
                    value = json.loads(line)
                except json.JSONDecodeError as exc:
                    yield {"__parse_error__": f"line {line_number}: {exc}"}
                    continue
                yield value if isinstance(value, dict) else {"__parse_error__": "non-object JSONL row"}


def iter_social(path: Path) -> Iterable[dict[str, Any]]:
    with path.open("r", encoding="utf-8") as handle:
        value = json.load(handle)
    if not isinstance(value, list):
        yield {"__parse_error__": "top-level social JSON is not an array"}
        return
    for item in value:
        if not isinstance(item, dict):
            yield {"__parse_error__": "non-object array item"}
            continue
        payload = item.get("identity_payload")
        if isinstance(payload, dict):
            # Provider identity attributes are intentionally nested in the
            # source JSON. Flatten only for verification and retain the raw
            # file unchanged.
            yield {**{key: value for key, value in item.items() if key != "identity_payload"}, **payload}
        else:
            yield item


def read_subscriptions(path: Path) -> tuple[list[str], Iterable[dict[str, Any]], list[str]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet_names = workbook.sheetnames
    sheet = workbook["subscriptions"] if "subscriptions" in workbook else workbook.worksheets[0]
    rows = sheet.iter_rows(values_only=True)
    title = next(rows, ())
    header = [str(value) if value is not None else "" for value in next(rows, ())]

    def records() -> Iterable[dict[str, Any]]:
        try:
            for row in rows:
                yield dict(zip(header, row))
        finally:
            workbook.close()

    return [str(value) if value is not None else "" for value in title], records(), sheet_names


def impossible_dob(value: Any) -> bool:
    if not isinstance(value, str) or not value:
        return False
    try:
        born = date.fromisoformat(value)
    except ValueError:
        return False
    today = date.today()
    age = today.year - born.year - ((today.month, today.day) < (born.month, born.day))
    return born > today or age < 5 or age > 110


def inspect_record(system: str, record: dict[str, Any], stats: SourceStats) -> None:
    stats.count += 1
    if "__parse_error__" in record or None in record:
        stats.malformed += 1
        return
    stats.fields.update(str(key) for key in record)
    for key in stats.fields:
        if key not in record or record.get(key) in (None, ""):
            stats.missing[key] += 1

    digest = fingerprint(record)
    if digest in stats._seen_hashes:
        stats.exact_extra += 1
    else:
        stats._seen_hashes.add(digest)

    id_field = ID_FIELDS.get(system)
    if id_field and record.get(id_field) not in (None, ""):
        record_id = str(record[id_field])
        known_hashes = stats._id_hashes[record_id]
        if known_hashes and digest not in known_hashes:
            stats.near_extra += 1
        known_hashes.add(digest)
        stats.ids.add(record_id)
        digits = re.sub(r"\D", "", record_id)
        if digits:
            numeric = int(digits)
            stats.numeric_ids.add(numeric)
            if stats._previous_numeric_id is not None and numeric < stats._previous_numeric_id:
                stats.order_decreased = True
            stats._previous_numeric_id = numeric

    joined = " ".join(str(value) for value in record.values() if value is not None)
    lowered = joined.lower()
    if "automation.internal" in lowered or re.search(r"\b(monitor|scraper|crawler|healthcheck|lb-probe)\w*", lowered):
        stats.bot_rows += 1
    if "@staff.test" in lowered or re.search(r"\b(?:qa|uat|load-test|staging)[+@\d]", lowered):
        stats.qa_rows += 1
    for feature, present in {
        "comma": "," in joined,
        "quote": '"' in joined,
        "line_break": "\n" in joined or "\r" in joined,
        "emoji": any(ord(char) > 0xFFFF for char in joined),
        "devanagari": any("\u0900" <= char <= "\u097f" for char in joined),
    }.items():
        if present:
            stats.text_features[feature] += 1

    for key, value in record.items():
        if value in (None, ""):
            continue
        key_lower = str(key).lower()
        text = str(value)
        if key_lower in {"city", "country", "device_type", "channel", "provider"}:
            stats.categories[key_lower].add(text)
        if key_lower in {"account_id", "app_account_ref", "user_id", "customer_id"}:
            stats.join_refs[key_lower].append(text)
        if "ts" in key_lower or key_lower.endswith("date"):
            fmt = timestamp_format(text)
            if fmt:
                stats.timestamp_formats[key_lower].add(fmt)
        if key_lower in {"phone", "contact_no"} and text in {"0000000000", "9999999999"}:
            stats.poison["placeholder_phone"] += 1
        if key_lower == "dob" and text in {"1900-01-01", "1970-01-01", "01-01-1900", "01-01-1970"}:
            stats.poison["default_dob"] += 1
        if "email" in key_lower and text.lower() == "bookings@events.example":
            stats.poison["corporate_email"] += 1
        if "email" in key_lower and text.lower().endswith("@staff.test"):
            stats.poison["test_email"] += 1
        if key_lower in {"device", "device_id"} and text == "KIOSK-DEVICE-1":
            stats.poison["kiosk_device"] += 1

    impossible = any(impossible_dob(record.get(key)) for key in ("dob", "date_of_birth"))
    if system == "ticketing":
        event = parse_timestamp(record.get("event_ts"))
        created = parse_timestamp(record.get("created_ts"))
        if event and created:
            stats.late_eligible += 1
            delta_days = (created - event).total_seconds() / 86_400
            if delta_days > 9:
                stats.late_rows += 1
            impossible = impossible or delta_days < 0
    if system == "subscriptions":
        start = parse_timestamp(record.get("start_date"))
        impossible = impossible or bool(start and start > datetime.now(timezone.utc))
        if record.get("billing_name") and record.get("billing_name") != record.get("subscriber_name"):
            stats.special["different_billing_name"] += 1
    if system == "ticketing" and record.get("guest") is True:
        stats.special["guest"] += 1
    if system == "social_logins" and record.get("provider"):
        provider = str(record["provider"]).lower()
        shape = tuple(sorted(str(key) for key in record if key not in {"provider", "provider_id"}))
        stats.provider_shapes[provider].add(shape)
    if impossible:
        stats.impossible_rows += 1
    elif any(
        "engagement" in str(key).lower()
        and isinstance(value, (int, float))
        and value >= 10_000
        for key, value in record.items()
    ):
        stats.impossible_rows += 1


def load_report(path: Path, audit: Audit) -> dict[str, Any]:
    if not path.exists():
        audit.add(False, "generation report", f"missing {path.name}")
        return {}
    try:
        with path.open("r", encoding="utf-8") as handle:
            value = json.load(handle)
    except (OSError, json.JSONDecodeError) as exc:
        audit.add(False, "generation report", f"cannot parse: {exc}")
        return {}
    audit.add(isinstance(value, dict), "generation report", "valid JSON object")
    return value if isinstance(value, dict) else {}


def find_numeric(document: Any, wanted: set[str]) -> float | None:
    if isinstance(document, dict):
        for key, value in document.items():
            if key in wanted and isinstance(value, (int, float)):
                return float(value)
            found = find_numeric(value, wanted)
            if found is not None:
                return found
    elif isinstance(document, list):
        for value in document:
            found = find_numeric(value, wanted)
            if found is not None:
                return found
    return None


def find_nested_dict(document: Any, wanted: str) -> dict[str, Any] | None:
    if isinstance(document, dict):
        value = document.get(wanted)
        if isinstance(value, dict):
            return value
        for child in document.values():
            found = find_nested_dict(child, wanted)
            if found is not None:
                return found
    elif isinstance(document, list):
        for child in document:
            found = find_nested_dict(child, wanted)
            if found is not None:
                return found
    return None


def verify(data_dir: Path) -> int:
    audit = Audit()
    stats: dict[str, SourceStats] = {}
    naive_clusters = NaiveClusterTracker()
    workbook_title: list[str] = []
    workbook_sheets: list[str] = []

    for system, filename in FILES.items():
        path = data_dir / filename
        if not path.exists():
            audit.add(False, f"{system} file", f"missing {filename}")
            continue
        try:
            if system in {"app_users", "store_customers"}:
                records = iter_csv(path)
            elif system == "ticketing":
                records = iter_jsonl(path)
            elif system == "social_logins":
                records = iter_social(path)
            else:
                workbook_title, records, workbook_sheets = read_subscriptions(path)
            source = SourceStats()
            for record in records:
                inspect_record(system, record, source)
                if "__parse_error__" not in record and None not in record:
                    naive_clusters.add(record)
            stats[system] = source
            audit.add(source.malformed == 0, f"{system} parsing", f"{source.count:,} rows; {source.malformed} malformed")
            missing_schema = EXPECTED_SCHEMAS[system] - source.fields
            audit.add(not missing_schema, f"{system} schema", f"missing fields: {sorted(missing_schema)}" if missing_schema else "required fields present")
        except Exception as exc:  # report all file failures in one run
            audit.add(False, f"{system} parsing", f"{type(exc).__name__}: {exc}")

    if "subscriptions" in stats:
        audit.add(len(workbook_sheets) > 1, "Excel sheets", f"sheets={workbook_sheets}")
        title_is_header = set(workbook_title) >= EXPECTED_SCHEMAS["subscriptions"]
        audit.add(bool(workbook_title) and not title_is_header, "Excel title row", f"row 1={workbook_title[:2]}")

    total_rows = sum(source.count for source in stats.values())
    audit.add(0 < total_rows < 3_000_000, "laptop row cap", f"{total_rows:,} rows")
    development_scale = total_rows / 420_000 if total_rows else 0
    scale_ok = 0 < total_rows < 100_000 or 350_000 <= total_rows <= 500_000
    audit.add(scale_ok, "roughly 420,000 records", f"{total_rows:,} observable rows (scale={development_scale:.4f})")

    mapping_path = data_dir / "person_map.csv"
    map_system_counts: Counter[str] = Counter()
    map_record_ids: dict[str, set[str]] = defaultdict(set)
    map_entity_types: Counter[str] = Counter()
    person_counts: Counter[str] = Counter()
    if mapping_path.exists():
        with mapping_path.open("r", encoding="utf-8", newline="") as handle:
            for row in csv.DictReader(handle):
                map_system_counts[row.get("system", "")] += 1
                map_entity_types[row.get("entity_type", "human")] += 1
                if row.get("record_id"):
                    map_record_ids[row.get("system", "")].add(row["record_id"])
                if row.get("person_id") and row.get("entity_type", "human") != "bot":
                    person_counts[row["person_id"]] += 1
        audit.add(True, "ground-truth map", f"{sum(map_system_counts.values()):,} mappings")
        coverage_ok = all(map_system_counts[name] == source.count for name, source in stats.items())
        detail = ", ".join(f"{name}: source={source.count:,}/map={map_system_counts[name]:,}" for name, source in stats.items())
        audit.add(coverage_ok, "ground-truth row coverage", detail)
        for name, source in stats.items():
            id_field = ID_FIELDS.get(name)
            addressable = bool(id_field) and source.ids == map_record_ids[name]
            audit.add(
                addressable,
                f"{name} addressable ground truth",
                f"source ID field={id_field!r}; source IDs={len(source.ids):,}; map IDs={len(map_record_ids[name]):,}",
            )
        distinct = len(person_counts)
        multi_rate = sum(count > 1 for count in person_counts.values()) / distinct if distinct else 0
        six_plus = sum(count >= 6 for count in person_counts.values())
        expected_people = 300_000 * development_scale
        audit.add(expected_people * 0.98 <= distinct <= expected_people * 1.02, "about 300,000 represented people", f"{distinct:,} (scaled target={expected_people:,.0f})")
        audit.add(0.20 <= multi_rate <= 0.30, "people with multiple records", f"{multi_rate:.2%}")
        audit.add(0 < six_plus <= max(1000, int(distinct * 0.01)), "small six-plus population", f"{six_plus:,} people")
    else:
        audit.add(False, "ground-truth map", "missing person_map.csv")

    exact = sum(source.exact_extra for source in stats.values())
    near = sum(source.near_extra for source in stats.values())
    exact_rate = exact / total_rows if total_rows else 0
    near_rate = near / total_rows if total_rows else 0
    audit.add(0.016 <= exact_rate <= 0.024, "exact duplicate rows", f"{exact:,} ({exact_rate:.2%})")
    audit.add(0.007 <= near_rate <= 0.013, "near-duplicate lower bound", f"{near:,} repeated IDs with changed content ({near_rate:.2%})")

    missing_rates: list[tuple[str, str, float]] = []
    for name, source in stats.items():
        if not source.count:
            continue
        for key, count in source.missing.items():
            rate = count / source.count
            if 0.04 <= rate <= 0.09:
                missing_rates.append((name, key, rate))
    audit.add(len(missing_rates) >= 3, "4%-9% missingness in three columns", str(missing_rates[:8]))
    email_rates = []
    for name, key in (("app_users", "email"), ("store_customers", "customer_email_address"), ("ticketing", "email")):
        if name in stats and stats[name].count:
            email_rates.append(stats[name].missing[key] / stats[name].count)
    audit.add(bool(email_rates) and max(email_rates) - min(email_rates) >= 0.02, "non-random/source-dependent missingness", f"email rates={[f'{rate:.2%}' for rate in email_rates]}")

    mixed_columns = [
        (name, key, sorted(formats))
        for name, source in stats.items()
        for key, formats in source.timestamp_formats.items()
        if len(formats) >= 4
    ]
    audit.add(bool(mixed_columns), "four timestamp formats in one column", str(mixed_columns[:4]))

    ticket = stats.get("ticketing")
    audit.add(bool(ticket and ticket.special["guest"]), "guest ticket bookings", f"{ticket.special['guest']:,}" if ticket else "ticketing missing")
    subscriptions = stats.get("subscriptions")
    billing_rate = subscriptions.special["different_billing_name"] / subscriptions.count if subscriptions and subscriptions.count else 0
    audit.add(billing_rate >= 0.40, "parent/partner billing names", f"{billing_rate:.2%} differ from subscriber")
    social = stats.get("social_logins")
    expected_providers = {"google", "facebook", "twitter", "apple"}
    providers = social.categories.get("provider", set()) if social else set()
    social_shapes_ok = bool(
        social
        and expected_providers == {value.lower() for value in providers}
        and any("hashed_email" in shape for shape in social.provider_shapes.get("apple", set()))
        and any("display_name" in shape for shape in social.provider_shapes.get("twitter", set()))
        and len({shape for shapes in social.provider_shapes.values() for shape in shapes}) >= 4
    )
    audit.add(social_shapes_ok, "four social-provider subsets", f"providers={sorted(providers)}; shapes={dict(social.provider_shapes) if social else {}}")

    all_fields = set().union(*(source.fields for source in stats.values())) if stats else set()
    for category in ("city", "country", "device_type", "channel"):
        values = set().union(*(source.categories.get(category, set()) for source in stats.values())) if stats else set()
        audit.add(category in all_fields and len(values) >= 3, f"{category} spelling variants", f"{len(values)} distinct values")

    if "app_users" in stats and "store_customers" in stats:
        store_ref_values = stats["store_customers"].join_refs.get("app_account_ref", [])
        store_refs = {
            int(re.sub(r"\D", "", value))
            for value in store_ref_values
            if re.sub(r"\D", "", value)
        }
        common_ids = stats["app_users"].numeric_ids & store_refs
        padded = any(value.startswith("0") for value in store_ref_values)
        audit.add(bool(common_ids) and padded, "misformatted shared join key", f"{len(common_ids):,} zero-padded store references overlap app integer IDs")
    if "ticketing" in stats:
        event_key = next((key for key in ("account_id", "user_id", "customer_id") if key in stats["ticketing"].fields), None)
        audit.add(event_key is not None, "event-table user join key", f"field={event_key!r}; needed to verify 3%-6% orphans")
        if event_key:
            app_numeric = stats.get("app_users", SourceStats()).numeric_ids
            references = stats["ticketing"].join_refs[event_key]
            orphan_count = sum(int(re.sub(r"\D", "", value) or -1) not in app_numeric for value in references)
            orphan_rate = orphan_count / len(references) if references else 0
            audit.add(0.03 <= orphan_rate <= 0.06, "event-table orphan IDs", f"{orphan_count:,}/{len(references):,} ({orphan_rate:.2%})")

    poison = Counter()
    for source in stats.values():
        poison.update(source.poison)

    def around_scaled(actual: int, full_target: int, tolerance: float = 0.20) -> bool:
        target = max(1, round(full_target * development_scale))
        return target * (1 - tolerance) <= actual <= target * (1 + tolerance)

    audit.add(around_scaled(poison["placeholder_phone"], 3_000), "placeholder phones", f"{poison['placeholder_phone']:,}")
    audit.add(around_scaled(poison["default_dob"], 4_200), "default DOBs", f"{poison['default_dob']:,}")
    audit.add(around_scaled(poison["corporate_email"], 900), "corporate booking email", f"{poison['corporate_email']:,}")
    audit.add(poison["kiosk_device"] >= max(10, round(30_000 * development_scale)), "shared kiosk device", f"{poison['kiosk_device']:,}")
    audit.add(around_scaled(poison["test_email"], 1_500), "staff test emails", f"{poison['test_email']:,}")

    late_rows = stats.get("ticketing", SourceStats()).late_rows
    late_total = stats.get("ticketing", SourceStats()).late_eligible
    late_rate = late_rows / late_total if late_total else 0
    audit.add(late_rate >= 0.03, "events arriving more than nine days late", f"{late_rows:,}/{late_total:,} ({late_rate:.2%})")
    ordered_sources = [name for name, source in stats.items() if name in ID_FIELDS and not source.order_decreased]
    audit.add(bool(stats) and not ordered_sources, "rows out of ID order", f"still monotonic: {ordered_sources}")

    bot_rows = map_entity_types["bot"] if map_entity_types else sum(source.bot_rows for source in stats.values())
    qa_rows = sum(source.qa_rows for source in stats.values())
    impossible_rows = sum(source.impossible_rows for source in stats.values())
    audit.add(total_rows > 0 and 0.04 <= bot_rows / total_rows <= 0.07, "automated traffic", f"{bot_rows:,} ({bot_rows / total_rows:.2%})" if total_rows else "no rows")
    audit.add(total_rows > 0 and 0.003 <= qa_rows / total_rows <= 0.007, "internal QA accounts", f"{qa_rows:,} ({qa_rows / total_rows:.2%})" if total_rows else "no rows")
    audit.add(total_rows > 0 and 0.002 <= impossible_rows / total_rows <= 0.004, "impossible values", f"{impossible_rows:,} ({impossible_rows / total_rows:.2%})" if total_rows else "no rows")

    features = Counter()
    for source in stats.values():
        features.update(source.text_features)
    for feature in ("comma", "quote", "line_break", "emoji", "devanagari"):
        audit.add(features[feature] > 0, f"free text: {feature}", f"{features[feature]:,} rows")

    hard_neg_path = data_dir / "hard_negatives.json"
    hard_negatives: Any = []
    if hard_neg_path.exists():
        try:
            hard_negatives = json.loads(hard_neg_path.read_text(encoding="utf-8"))
        except json.JSONDecodeError:
            hard_negatives = []
    types = {str(item.get("type", "")).lower() for item in hard_negatives if isinstance(item, dict)} if isinstance(hard_negatives, list) else set()
    required_hard_negatives = {
        "father/son": any("father" in item or "parent" in item for item in types),
        "university lab device": any("university" in item or "computer_lab" in item for item in types),
        "common name/city": any("common_name" in item for item in types),
        "couple shared email+token": any("couple" in item and "email" in item for item in types),
    }
    audit.add(all(required_hard_negatives.values()), "required hard-negative scenarios", str(required_hard_negatives))

    report = load_report(data_dir / "generation_report.json", audit)
    measured_hn = find_numeric(report, {
        "measured_hard_negative_candidate_pair_rate",
        "hard_negative_candidate_pair_rate",
        "explicit_hard_negative_rate_rule2",
    })
    if measured_hn is None:
        audit.warn("hard negatives >=5% of naive candidates", "no measured rate or documented candidate definition in report")
    else:
        audit.add(measured_hn >= 0.05, "hard negatives >=5% of naive candidates", f"{measured_hn:.2%}")
    measured_zero = find_numeric(report, {"measured_zero_evidence_duplicate_pair_rate", "zero_evidence_duplicate_pair_rate"})
    if measured_zero is None:
        audit.warn("8% zero-evidence true pairs", "no measured pair-level rate in report")
    else:
        audit.add(0.07 <= measured_zero <= 0.09, "8% zero-evidence true pairs", f"{measured_zero:.2%}")
    mode_counts = find_nested_dict(report, "duplicate_evidence_mode_counts")
    mode_groups = {
        "exact email": {"exact_email", "exact_verified_email"},
        "email case/dot/plus": {"email_variant", "email_case_variation", "email_dotted_local_part", "email_plus_suffix"},
        "phone formatting": {"phone_variant", "phone_country_code", "phone_spaced"},
        "name+city only": {"name_city", "name_city_only"},
        "device only": {"device_only"},
        "zero evidence": {"zero", "no_usable_evidence"},
    }
    covered_modes = {
        group: any((mode_counts or {}).get(alias, 0) > 0 for alias in aliases)
        for group, aliases in mode_groups.items()
    }
    audit.add(
        bool(mode_counts) and all(covered_modes.values()),
        "duplicate evidence-quality mix",
        f"covered={covered_modes}; mode counts={mode_counts}",
    )
    source_backed = bool(hard_negatives) and all(
        isinstance(item, dict)
        and len(item.get("person_ids", [])) == 2
        and len(set(item.get("person_ids", []))) == 2
        and len(item.get("source_records", [])) == 2
        for item in hard_negatives
    )
    audit.add(source_backed, "hard negatives are source-backed", f"{len(hard_negatives) if isinstance(hard_negatives, list) else 0} manifest entries")

    reported_counts = report.get("row_counts", {}) if isinstance(report, dict) else {}
    if isinstance(reported_counts, dict) and stats:
        consistent = all(reported_counts.get(name) == source.count for name, source in stats.items())
        audit.add(consistent, "reported row counts are measured", f"reported={reported_counts}")
    if person_counts:
        actual_distinct = len(person_counts)
        actual_multi = sum(count > 1 for count in person_counts.values()) / actual_distinct
        reported_distinct = report.get("true_number_of_distinct_people") if isinstance(report, dict) else None
        reported_multi = report.get("true_duplicate_rate_people_with_multiple_records") if isinstance(report, dict) else None
        audit.add(reported_distinct == actual_distinct, "reported distinct-person count", f"reported={reported_distinct}, measured={actual_distinct}")
        audit.add(
            isinstance(reported_multi, (int, float)) and math.isclose(float(reported_multi), actual_multi, abs_tol=0.0001),
            "reported true duplicate rate",
            f"reported={reported_multi}, measured={actual_multi:.4f}",
        )
    largest_actual = naive_clusters.largest_poisoned_cluster()
    largest_reported = None
    if isinstance(report, dict):
        largest_reported = report.get("largest_poisoned_cluster_naive_matcher", report.get("estimated_largest_poisoned_cluster"))
    audit.add(
        largest_reported == largest_actual and largest_actual >= max(poison.values(), default=0),
        "largest poisoned transitive cluster report",
        f"reported={largest_reported}, independently observed={largest_actual}",
    )

    definitions = report.get("definitions", {}) if isinstance(report, dict) else {}
    audit.add(bool(definitions.get("timestamp_semantics")), "local-time versus UTC semantics", str(definitions.get("timestamp_semantics")))
    audit.add(bool(definitions.get("reserved_phone_range")), "reserved phone range", str(definitions.get("reserved_phone_range")))
    return audit.print()


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--data-dir",
        type=Path,
        default=Path(__file__).resolve().parents[1] / "data" / "generated",
        help="directory containing the generated files",
    )
    args = parser.parse_args()
    return verify(args.data_dir.resolve())


if __name__ == "__main__":
    sys.exit(main())
