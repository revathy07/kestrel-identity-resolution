"""Read the five normal source systems without accessing evaluation truth.

Readers return the original field names and values. Dotted field paths in the schema
mapping are resolved only when a caller asks for a value; nested JSON is not flattened or
rewritten in the raw record.
"""

from __future__ import annotations

import csv
import json
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterator, Mapping

import openpyxl


class SourceReadError(RuntimeError):
    """Raised when a normal source file is missing or violates its declared schema."""


@dataclass(frozen=True)
class SourceRecord:
    source: str
    source_record_id: str
    raw: Mapping[str, Any]


def load_schema_mapping(path: Path) -> dict[str, Any]:
    """Load the JSON-compatible YAML mapping and validate its outer structure."""

    try:
        with path.open("r", encoding="utf-8") as handle:
            mapping = json.load(handle)
    except FileNotFoundError as exc:
        raise SourceReadError(f"Schema mapping not found: {path}") from exc
    except json.JSONDecodeError as exc:
        raise SourceReadError(f"Invalid schema mapping {path}: {exc}") from exc

    if not isinstance(mapping, dict) or not isinstance(mapping.get("sources"), dict):
        raise SourceReadError("Schema mapping must contain a 'sources' object")
    if not isinstance(mapping.get("canonical_concepts"), list):
        raise SourceReadError("Schema mapping must contain a 'canonical_concepts' list")
    return mapping


def get_raw_value(record: Mapping[str, Any], dotted_path: str) -> Any:
    """Resolve a dotted path without modifying the original mapping."""

    value: Any = record
    for part in dotted_path.split("."):
        if not isinstance(value, Mapping) or part not in value:
            return None
        value = value[part]
    return value


def _record_id(raw: Mapping[str, Any], record_id_path: str, source: str, ordinal: int) -> str:
    value = get_raw_value(raw, record_id_path)
    if value is None or not str(value).strip():
        raise SourceReadError(
            f"{source} row {ordinal:,} has no source record ID at {record_id_path!r}"
        )
    return str(value)


def _validate_columns(
    columns: list[str], required_paths: list[str], source: str, *, nested: bool = False
) -> None:
    if nested:
        return
    missing = sorted(path for path in required_paths if "." not in path and path not in columns)
    if missing:
        raise SourceReadError(f"{source} is missing mapped columns: {', '.join(missing)}")


def _iter_csv(path: Path, source: str, spec: Mapping[str, Any]) -> Iterator[SourceRecord]:
    try:
        handle = path.open("r", encoding="utf-8-sig", newline="")
    except FileNotFoundError as exc:
        raise SourceReadError(f"Required source file not found: {path}") from exc
    with handle:
        reader = csv.DictReader(handle)
        if not reader.fieldnames:
            raise SourceReadError(f"{source} has no CSV header")
        required = [spec["record_id"], *spec["identifiers"].keys()]
        _validate_columns(reader.fieldnames, required, source)
        for ordinal, raw in enumerate(reader, start=1):
            yield SourceRecord(source, _record_id(raw, spec["record_id"], source, ordinal), raw)


def _iter_jsonl(path: Path, source: str, spec: Mapping[str, Any]) -> Iterator[SourceRecord]:
    try:
        handle = path.open("r", encoding="utf-8")
    except FileNotFoundError as exc:
        raise SourceReadError(f"Required source file not found: {path}") from exc
    with handle:
        for ordinal, line in enumerate(handle, start=1):
            if not line.strip():
                continue
            try:
                raw = json.loads(line)
            except json.JSONDecodeError as exc:
                raise SourceReadError(f"Invalid JSON in {source} line {ordinal}: {exc}") from exc
            if not isinstance(raw, dict):
                raise SourceReadError(f"{source} line {ordinal} is not a JSON object")
            if ordinal == 1:
                required = [spec["record_id"], *spec["identifiers"].keys()]
                _validate_columns(list(raw), required, source)
            yield SourceRecord(source, _record_id(raw, spec["record_id"], source, ordinal), raw)


def _excel_header(
    workbook: openpyxl.Workbook, source: str, spec: Mapping[str, Any]
) -> tuple[Any, int, list[str]]:
    marker = str(spec.get("header_marker", spec["record_id"]))
    preferred_sheet = spec.get("sheet")
    sheets = (
        [workbook[preferred_sheet]]
        if preferred_sheet and preferred_sheet in workbook.sheetnames
        else list(workbook.worksheets)
    )
    for sheet in sheets:
        for row_number, row in enumerate(
            sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 25), values_only=True), start=1
        ):
            values = ["" if value is None else str(value) for value in row]
            if marker in values:
                headers = values
                required = [spec["record_id"], *spec["identifiers"].keys()]
                _validate_columns(headers, required, source)
                return sheet, row_number, headers
    raise SourceReadError(
        f"Could not locate Excel header marker {marker!r} for {source} in the first 25 rows"
    )


def _iter_excel(path: Path, source: str, spec: Mapping[str, Any]) -> Iterator[SourceRecord]:
    if not path.exists():
        raise SourceReadError(f"Required source file not found: {path}")
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        sheet, header_row, headers = _excel_header(workbook, source, spec)
        for ordinal, values in enumerate(
            sheet.iter_rows(min_row=header_row + 1, values_only=True), start=1
        ):
            if all(value is None for value in values):
                continue
            raw = {header: value for header, value in zip(headers, values) if header}
            yield SourceRecord(source, _record_id(raw, spec["record_id"], source, ordinal), raw)
    finally:
        workbook.close()


def _iter_nested_json(path: Path, source: str, spec: Mapping[str, Any]) -> Iterator[SourceRecord]:
    try:
        with path.open("r", encoding="utf-8") as handle:
            payload = json.load(handle)
    except FileNotFoundError as exc:
        raise SourceReadError(f"Required source file not found: {path}") from exc
    except json.JSONDecodeError as exc:
        raise SourceReadError(f"Invalid JSON in {source}: {exc}") from exc
    if not isinstance(payload, list):
        raise SourceReadError(f"{source} must contain a JSON array")
    required_paths = [spec["record_id"], *spec["identifiers"].keys()]
    for ordinal, raw in enumerate(payload, start=1):
        if not isinstance(raw, dict):
            raise SourceReadError(f"{source} array item {ordinal} is not an object")
        identity_payload = raw.get("identity_payload")
        if not isinstance(identity_payload, dict):
            raise SourceReadError(f"{source} array item {ordinal} has no identity_payload object")
        if ordinal == 1:
            _validate_columns(list(raw), required_paths, source, nested=True)
        yield SourceRecord(source, _record_id(raw, spec["record_id"], source, ordinal), raw)


READERS = {
    "csv": _iter_csv,
    "jsonl": _iter_jsonl,
    "excel": _iter_excel,
    "nested_json": _iter_nested_json,
}


def iter_source_records(
    data_dir: Path, source: str, spec: Mapping[str, Any]
) -> Iterator[SourceRecord]:
    """Yield raw records for one declared normal source."""

    source_format = spec.get("format")
    if source_format not in READERS:
        raise SourceReadError(f"Unsupported format {source_format!r} for {source}")
    relative_path = spec.get("path")
    if not isinstance(relative_path, str) or not relative_path:
        raise SourceReadError(f"No relative path configured for {source}")
    path = data_dir / relative_path
    yield from READERS[source_format](path, source, spec)
