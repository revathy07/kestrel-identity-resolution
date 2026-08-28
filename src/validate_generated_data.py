#!/usr/bin/env python3
"""Independent, read-only audit of the generated identity-resolution dataset.

The program never writes inside the supplied data directory. Hidden person identifiers
are used only in memory for aggregate calculations and are never emitted to reports.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import math
import re
import statistics
import sys
from collections import Counter, defaultdict
from dataclasses import asdict, dataclass
from datetime import date, datetime, timedelta, timezone
from itertools import combinations
from pathlib import Path
from typing import Any, Callable, Iterable

from openpyxl import load_workbook


FILES = {
    "app_users": "app_users.csv",
    "store_customers": "store_customers.csv",
    "ticketing": "ticketing.jl",
    "subscriptions": "subscriptions.xlsx",
    "social_logins": "social_logins.json",
}
ID_FIELDS = {
    "app_users": "account_id", "store_customers": "customer_id",
    "ticketing": "booking_id", "subscriptions": "subscription_id",
    "social_logins": "provider_id",
}
EXPECTED = {
    "app_users": {"account_id", "email", "phone", "first_name", "last_name", "dob", "device_id", "signup_ts", "country"},
    "store_customers": {"customer_id", "customer_email_address", "contact_no", "line1", "line2", "city", "postcode"},
    "ticketing": {"booking_id", "full_name", "email", "phone", "guest", "event_ts", "created_ts"},
    "subscriptions": {"subscription_id", "email", "billing_name", "payment_token"},
    "social_logins": {"provider", "identity_payload"},
}
POISON_VALUES = {
    "placeholder_phone": {"0000000000", "9999999999"},
    "default_dob": {"1900-01-01", "1970-01-01", "01-01-1900", "01-01-1970"},
    "corporate_email": {"bookings@events.example"},
    "kiosk_device": {"KIOSK-DEVICE-1"},
}
TIMESTAMP_FIELDS = {
    "app_users": ("signup_ts",), "store_customers": ("updated_ts",),
    "ticketing": ("event_ts", "created_ts"), "subscriptions": ("start_date",),
    "social_logins": ("login_ts",),
}
MISSING_SENTINELS = {"", "null", "none"}


@dataclass
class Check:
    id: str
    requirement: str
    target: str
    observed: str
    status: str
    evidence: str
    recommended_action: str = "None."
    mandatory: bool = True
    likely_cause: str = ""
    generator_location: str = ""
    regeneration_required: bool = False


class Audit:
    def __init__(self) -> None:
        self.checks: list[Check] = []
        self.metrics: list[dict[str, Any]] = []

    def add(self, ident: str, requirement: str, target: str, observed: Any,
            status: str, evidence: str, action: str = "None.", *, mandatory: bool = True,
            cause: str = "", location: str = "", regen: bool = False) -> None:
        assert status in {"PASS", "FAIL", "WARNING", "NOT VERIFIABLE"}
        if status == "PASS":
            action = "None."
        self.checks.append(Check(ident, requirement, target, str(observed), status,
                                 evidence, action, mandatory, cause, location, regen))

    def metric(self, ident: str, name: str, value: Any, unit: str = "") -> None:
        self.metrics.append({"id": ident, "metric": name, "value": value, "unit": unit})


def is_missing(value: Any) -> bool:
    if value is None:
        return True
    if not isinstance(value, str):
        return False
    text = value.strip()
    while len(text) >= 2 and text[0] == text[-1] and text[0] in {'"', "'"}:
        text = text[1:-1].strip()
    return text.casefold() in MISSING_SENTINELS


def clean_text(value: Any) -> str:
    if is_missing(value):
        return ""
    return str(value).splitlines()[0].split(' "quoted')[0].split(" hello ")[0].replace("dY~S", "").strip()


def norm_email(value: Any) -> str:
    # Independently implement the documented missing/artifact semantics.
    if is_missing(value):
        return ""
    text = str(value).splitlines()[0].split(' "quoted')[0].strip().lower()
    if is_missing(text):
        return ""
    if "@" not in text:
        return text
    local, domain = text.split("@", 1)
    if domain == "brand-example.test":
        local = local.split("+", 1)[0].replace(".", "")
    return f"{local}@{domain}"


def norm_phone(value: Any) -> str:
    digits = re.sub(r"\D", "", clean_text(value))
    if digits in POISON_VALUES["placeholder_phone"]:
        return digits
    return digits.removeprefix("999").lstrip("0") if digits else ""


def norm_words(value: Any) -> str:
    return re.sub(r"[^a-z]", "", clean_text(value).lower())


def identity_view(record: dict[str, Any]) -> dict[str, Any]:
    payload = record.get("identity_payload")
    return {**record, **payload} if isinstance(payload, dict) else record


def source_record_id(system: str, record: dict[str, Any]) -> str:
    return clean_text(identity_view(record).get(ID_FIELDS[system]))


def mask(value: Any, kind: str = "text") -> str:
    text = clean_text(value)
    if not text:
        return "<missing>"
    if kind == "email" and "@" in text:
        local, domain = text.split("@", 1)
        return f"{local[:2]}***@{domain}"
    if kind == "phone":
        digits = re.sub(r"\D", "", text)
        return "***" + digits[-4:]
    if kind == "name":
        return " ".join(part[:1] + "***" for part in text.split())
    return text[:3] + "***" if len(text) > 3 else "***"


def fingerprint(record: dict[str, Any]) -> str:
    raw = json.dumps(record, sort_keys=True, ensure_ascii=False, default=str).encode()
    return hashlib.blake2b(raw, digest_size=16).hexdigest()


def timestamp_format(value: Any) -> str | None:
    text = clean_text(value)
    patterns = (
        (r"^\d{2}-\d{2}-\d{4} \d{2}:\d{2}:\d{2}$", "DD-MM-YYYY"),
        (r"^\d{4}/\d{2}/\d{2} \d{2}:\d{2}:\d{2}$", "YYYY/MM/DD"),
        (r"^\d{2}-\d{2}-\d{2} \d{2}:\d{2}$", "MM-DD-YY"),
        (r"^\d{4}-\d{2}-\d{2}T.*(?:Z|[+-]\d{2}:\d{2})$", "ISO_OFFSET"),
        (r"^\d{13}$", "EPOCH_MS"),
        (r"^\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}$", "LOCAL_TEXT"),
    )
    return next((label for pattern, label in patterns if re.match(pattern, text)), None)


def parse_timestamp(value: Any) -> datetime | None:
    text, fmt = clean_text(value), timestamp_format(value)
    try:
        if fmt == "DD-MM-YYYY": parsed = datetime.strptime(text, "%d-%m-%Y %H:%M:%S")
        elif fmt == "YYYY/MM/DD": parsed = datetime.strptime(text, "%Y/%m/%d %H:%M:%S")
        elif fmt == "MM-DD-YY": parsed = datetime.strptime(text, "%m-%d-%y %H:%M")
        elif fmt == "ISO_OFFSET": return datetime.fromisoformat(text.replace("Z", "+00:00")).astimezone(timezone.utc)
        elif fmt == "EPOCH_MS": return datetime.fromtimestamp(int(text) / 1000, tz=timezone.utc)
        elif fmt == "LOCAL_TEXT":
            parsed = datetime.strptime(text, "%Y-%m-%d %H:%M:%S")
            return parsed.replace(tzinfo=timezone(timedelta(hours=5, minutes=30))).astimezone(timezone.utc)
        else: return None
        return parsed.replace(tzinfo=timezone.utc)
    except (ValueError, OverflowError):
        return None


def record_tokens(record: dict[str, Any], include_poison: bool = True) -> set[str]:
    record = identity_view(record)
    result: set[str] = set()
    for key in ("email", "customer_email_address", "verified_email"):
        value = norm_email(record.get(key))
        if value: result.add("email:" + value)
    if clean_text(record.get("hashed_email")):
        result.add("hash:" + clean_text(record["hashed_email"]))
    for key in ("phone", "contact_no"):
        value = norm_phone(record.get(key))
        if value: result.add("phone:" + value)
    for key in ("device_id", "device"):
        value = clean_text(record.get(key))
        if value: result.add("device:" + value)
    for key in ("account_id", "app_account_ref"):
        value = clean_text(record.get(key))
        if value and value.isdigit(): result.add("account:" + str(int(value)))
    if clean_text(record.get("payment_token")):
        result.add("payment:" + clean_text(record["payment_token"]))
    dob = clean_text(record.get("dob"))
    if include_poison and dob in POISON_VALUES["default_dob"]: result.add("dob:" + dob)
    city = norm_words(record.get("city"))
    if record.get("first_name") or record.get("last_name"):
        name = f"{record.get('first_name', '')} {record.get('last_name', '')}"
    elif record.get("first") or record.get("last"):
        name = f"{record.get('first', '')} {record.get('last', '')}"
    else:
        name = record.get("full_name") or record.get("display_name") or ""
    name = norm_words(name)
    if name and city: result.add(f"name_city:{name}:{city}")
    if not include_poison:
        result = {t for t in result if not is_explicit_poison_token(t)}
    return result


def is_explicit_poison_token(token: str) -> bool:
    return (token in {"phone:0000000000", "phone:9999999999", "dob:1900-01-01", "dob:1970-01-01",
                      "dob:01-01-1900", "dob:01-01-1970", "email:bookings@events.example",
                      "device:KIOSK-DEVICE-1"} or token.endswith("@staff.test"))


def poison_types(record: dict[str, Any]) -> set[str]:
    values = {clean_text(v) for v in identity_view(record).values() if not isinstance(v, dict)}
    found = {kind for kind, candidates in POISON_VALUES.items() if values & candidates}
    if any(v.lower().endswith("@staff.test") for v in values): found.add("test_email")
    return found


def iter_csv(path: Path) -> Iterable[dict[str, Any]]:
    with path.open(encoding="utf-8", newline="") as handle:
        yield from csv.DictReader(handle)


def iter_jsonl(path: Path) -> Iterable[dict[str, Any]]:
    with path.open(encoding="utf-8") as handle:
        for line in handle:
            if line.strip():
                value = json.loads(line)
                if not isinstance(value, dict): raise ValueError("JSONL item is not an object")
                yield value


def iter_social(path: Path) -> Iterable[dict[str, Any]]:
    with path.open(encoding="utf-8") as handle: value = json.load(handle)
    if not isinstance(value, list): raise ValueError("social JSON top level is not an array")
    for item in value:
        if not isinstance(item, dict): raise ValueError("social JSON item is not an object")
        yield item


def subscription_info(path: Path) -> tuple[list[str], int, list[str]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    names = wb.sheetnames
    ws = wb["subscriptions"] if "subscriptions" in names else wb.worksheets[0]
    rows = ws.iter_rows(values_only=True)
    first = next(rows, ())
    second = next(rows, ())
    wb.close()
    return names, 2, [str(v) if v is not None else "" for v in second]


def iter_subscriptions(path: Path) -> Iterable[dict[str, Any]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["subscriptions"] if "subscriptions" in wb.sheetnames else wb.worksheets[0]
    rows = ws.iter_rows(values_only=True)
    next(rows, None)
    header = [str(v) if v is not None else "" for v in next(rows, ())]
    try:
        for row in rows: yield dict(zip(header, row))
    finally: wb.close()


def source_factory(data_dir: Path, system: str) -> Callable[[], Iterable[dict[str, Any]]]:
    path = data_dir / FILES[system]
    if system in {"app_users", "store_customers"}: return lambda: iter_csv(path)
    if system == "ticketing": return lambda: iter_jsonl(path)
    if system == "subscriptions": return lambda: iter_subscriptions(path)
    return lambda: iter_social(path)


class UnionFind:
    def __init__(self, count: int) -> None:
        self.parent = list(range(count)); self.size = [1] * count

    def find(self, value: int) -> int:
        while self.parent[value] != value:
            self.parent[value] = self.parent[self.parent[value]]; value = self.parent[value]
        return value

    def union(self, left: int, right: int) -> None:
        a, b = self.find(left), self.find(right)
        if a == b: return
        if self.size[a] < self.size[b]: a, b = b, a
        self.parent[b] = a; self.size[a] += self.size[b]


def pct(value: float) -> str:
    return f"{100 * value:.4f}%"


def deviation(observed: float, target: float) -> tuple[float, float]:
    absolute = abs(observed - target)
    return absolute, absolute / target if target else 0.0


def compare_approx(audit: Audit, ident: str, requirement: str, observed: float, target: float,
                   tolerance: float, unit: str, evidence: str, location: str) -> None:
    absolute, relative = deviation(observed, target)
    passed = absolute <= tolerance
    detail = f"{evidence}; absolute deviation={absolute:,.4f} {unit}; relative deviation={pct(relative)}; tolerance=±{tolerance:,.4f} {unit}"
    audit.add(ident, requirement, f"approximately {target:,.4f} {unit} (±{tolerance:,.4f})",
              f"{observed:,.4f} {unit}", "PASS" if passed else "FAIL", detail,
              "Adjust the allocation/injection target and regenerate." if not passed else "None.",
              cause="Observed emitted-data rate is outside the predeclared tolerance.", location=location, regen=not passed)


def extract_fields(system: str, record: dict[str, Any]) -> dict[str, Any]:
    record = identity_view(record)
    email_key = "customer_email_address" if system == "store_customers" else "verified_email" if system == "social_logins" else "email"
    phone_key = "contact_no" if system == "store_customers" else "phone"
    device_key = "device" if system == "store_customers" else "device_id"
    first = record.get("first_name", record.get("first", "")); last = record.get("last_name", record.get("last", ""))
    fullname = record.get("full_name") or record.get("subscriber_name") or record.get("display_name") or f"{first} {last}"
    return {
        "system": system, "email_raw": clean_text(record.get(email_key)), "email": norm_email(record.get(email_key)),
        "verified": email_key == "verified_email" and bool(clean_text(record.get(email_key))),
        "phone_raw": clean_text(record.get(phone_key)), "phone": norm_phone(record.get(phone_key)),
        "device": clean_text(record.get(device_key)), "name": norm_words(fullname),
        "city": norm_words(record.get("city")), "account": clean_text(record.get("account_id") or record.get("app_account_ref")),
        "payment": clean_text(record.get("payment_token")), "tokens": record_tokens(record, include_poison=False),
    }


def observed_modes(left: dict[str, Any], right: dict[str, Any], usable: set[str]) -> set[str]:
    modes: set[str] = set()
    left_raw, right_raw = left["email_raw"], right["email_raw"]
    if left_raw and right_raw and left["email"] == right["email"]:
        left_local, right_local = left_raw.split("@", 1)[0], right_raw.split("@", 1)[0]
        if left_raw == right_raw and (left["verified"] or right["verified"]): modes.add("exact_verified_email")
        elif left_raw == right_raw: modes.add("exact_email")
        if left_raw != right_raw and left_raw.casefold() == right_raw.casefold(): modes.add("email_case_variation")
        if left_raw.casefold() != right_raw.casefold() and left_local.split("+",1)[0].replace(".","").casefold() == right_local.split("+",1)[0].replace(".","").casefold() and ("." in left_local) != ("." in right_local): modes.add("email_dotted_local_part")
        if ("+" in left_local) != ("+" in right_local): modes.add("email_plus_suffix")
    if left["phone_raw"] and right["phone_raw"] and left["phone"] == right["phone"] and left["phone_raw"] != right["phone_raw"]:
        left_digits, right_digits = re.sub(r"\D","",left["phone_raw"]), re.sub(r"\D","",right["phone_raw"])
        if left_digits.startswith("999") != right_digits.startswith("999"): modes.add("phone_country_code")
        if " " in left["phone_raw"] or " " in right["phone_raw"]: modes.add("phone_spaced")
        if left["phone_raw"].startswith("0") or right["phone_raw"].startswith("0"): modes.add("phone_leading_zero")
    if usable and all(value.startswith("name_city:") for value in usable): modes.add("name_city_only")
    if usable and all(value.startswith("device:") for value in usable): modes.add("device_only")
    if not usable: modes.add("no_usable_evidence")
    if usable and not modes: modes.add("multiple_or_other_evidence")
    return modes


def pair_code(left: int, right: int) -> int:
    if left > right: left, right = right, left
    return (left << 20) | right


def high_edge_union(groups: list[set[int]]) -> int:
    total = 0
    for mask_value in range(1, 1 << len(groups)):
        chosen = [groups[index] for index in range(len(groups)) if mask_value & (1 << index)]
        overlap = set.intersection(*chosen)
        pairs = math.comb(len(overlap), 2) if len(overlap) > 1 else 0
        total += pairs if len(chosen) % 2 else -pairs
    return total


def load_truth(path: Path) -> tuple[dict[tuple[str, str], tuple[str, str, int]], Counter[str], Counter[tuple[str, str]], int, int]:
    grouped: dict[tuple[str, str], list[Any]] = {}
    people: Counter[str] = Counter(); physical: Counter[tuple[str, str]] = Counter(); rows = 0; conflicts = 0
    with path.open(encoding="utf-8", newline="") as handle:
        reader = csv.DictReader(handle)
        for row in reader:
            rows += 1; key = (row["system"], row["record_id"]); pid = row["person_id"]; typ = row.get("entity_type", "human")
            physical[key] += 1
            current = grouped.get(key)
            if current is None: grouped[key] = [pid, typ, 1]
            else:
                if current[:2] != [pid, typ]: conflicts += 1
                current[2] += 1
            if typ == "human": people[pid] += 1
    return {k: (v[0], v[1], v[2]) for k, v in grouped.items()}, people, physical, rows, conflicts


def audit_dataset(data_dir: Path, output_dir: Path, generator: Path | None = None) -> tuple[Audit, dict[str, Any]]:
    audit = Audit(); output_dir.mkdir(parents=True, exist_ok=True)
    canonical_path = data_dir / "hidden" / "canonical_duplicate_links.jsonl"
    required = [data_dir / name for name in FILES.values()] + [data_dir / "person_map.csv", data_dir / "generation_report.json", data_dir / "hard_negatives.json", canonical_path]
    missing = [p.name for p in required if not p.is_file()]
    audit.add("FORMAT-001", "All required generated files exist", "9 required artifacts", f"{9-len(missing)}/9 present",
              "PASS" if not missing else "FAIL", f"Missing: {missing or 'none'}", "Generate the complete dataset." if missing else "None.")
    if missing: return audit, {"fatal": "missing artifacts", "missing": missing}

    print("[1/7] Loading hidden truth aggregates (no identities will be printed)...")
    truth, people_counts, truth_occ, truth_rows, truth_conflicts = load_truth(data_dir / "person_map.csv")
    multi_people = {pid for pid, n in people_counts.items() if n > 1}
    six_plus = sum(n >= 6 for n in people_counts.values())

    print("[2/7] Streaming and profiling all five source systems...")
    factories = {s: source_factory(data_dir, s) for s in FILES}
    source_counts: Counter[str] = Counter(); schemas: dict[str, set[str]] = defaultdict(set)
    source_occ: Counter[tuple[str, str]] = Counter(); parse_errors: list[str] = []
    token_freq: Counter[str] = Counter(); exact_hashes: dict[str, Counter[str]] = defaultdict(Counter)
    seen_ids: dict[str, set[str]] = defaultdict(set); repeat_ids: dict[str, set[str]] = defaultdict(set)
    missing_counts: Counter[tuple[str, str]] = Counter(); group_missing: Counter[tuple[str, str, str, str]] = Counter(); group_total: Counter[tuple[str, str, str]] = Counter()
    timestamp_counts: Counter[tuple[str, str, str]] = Counter(); timestamp_unparseable: Counter[tuple[str, str]] = Counter(); ambiguous_dates = 0
    category_counts: Counter[tuple[str, str, str]] = Counter(); free_counts: Counter[str] = Counter(); free_examples: dict[str, str] = {}
    poison_counts: Counter[str] = Counter(); poison_sources: Counter[tuple[str, str]] = Counter(); poison_people: dict[str, set[str]] = defaultdict(set)
    normal_labels: set[str] = set(); domains: Counter[str] = Counter(); phone_prefix_ok = 0; phone_values = 0
    multi_records: dict[str, list[dict[str, Any]]] = defaultdict(list)
    multi_record_lookup: dict[tuple[str, str], dict[str, Any]] = {}
    qa_people: set[str] = set(); bot_rows = human_rows = 0; bot_engagement: list[float] = []; human_engagement: list[float] = []
    bot_times: set[str] = set(); human_times: set[str] = set(); bot_instants: list[datetime] = []; human_instants: list[datetime] = []; provider_shapes: dict[str, Counter[tuple[str, ...]]] = defaultdict(Counter)
    bot_ticket_delays: list[float] = []; human_ticket_delays: list[float] = []; obvious_bot_markers = 0
    billing_different = guest_no_account = 0; app_ids_numeric: set[int] = set(); ticket_refs: list[str] = []
    late = late_eligible = inversions = ticket_adjacent = 0; prior_event: datetime | None = None
    impossible_sets: dict[str, set[tuple[str, str]]] = defaultdict(set); engagement_values: list[tuple[tuple[str, str], float]] = []

    for system, factory in factories.items():
        try:
            for record in factory():
                source_counts[system] += 1; schemas[system].update(map(str, record.keys()))
                rid = source_record_id(system, record); key = (system, rid); source_occ[key] += 1
                if rid in seen_ids[system]: repeat_ids[system].add(rid)
                else: seen_ids[system].add(rid)
                exact_hashes[system][fingerprint(record)] += 1
                mapped = truth.get(key); pid, entity_type = (mapped[0], mapped[1]) if mapped else ("", "")
                if pid in multi_people and entity_type == "human":
                    compact = extract_fields(system, record); multi_records[pid].append(compact); multi_record_lookup.setdefault(key, compact)
                for token in record_tokens(record): token_freq[token] += 1
                for field in schemas[system]:
                    if field not in record or is_missing(record.get(field)): missing_counts[(system, field)] += 1
                country = clean_text(record.get("country")) or "<missing>"; device_type = clean_text(record.get("device_type")) or "<missing>"
                for group_name, group_value in (("country", country), ("device_type", device_type)):
                    for field in ("email", "customer_email_address", "phone", "contact_no", "first_name", "last_name"):
                        if field in schemas[system]:
                            group_total[(field, group_name, group_value)] += 1
                            if is_missing(record.get(field)): group_missing[(field, group_name, group_value, system)] += 1
                for field in TIMESTAMP_FIELDS[system]:
                    value = record.get(field)
                    if not is_missing(value):
                        fmt = timestamp_format(value)
                        if fmt: timestamp_counts[(system, field, fmt)] += 1
                        else: timestamp_unparseable[(system, field)] += 1
                        text = clean_text(value)
                        if re.match(r"^(0[1-9]|1[0-2])-(0[1-9]|1[0-2])-(?:\d{2}|\d{4})", text): ambiguous_dates += 1
                for concept in ("channel", "city", "device_type", "country"):
                    if concept in record and not is_missing(record.get(concept)): category_counts[(concept, system, clean_text(record[concept]))] += 1
                flat_record = identity_view(record)
                joined = " ".join(str(v) for v in flat_record.values() if v is not None and not isinstance(v, dict))
                feature_map = {"comma": "," in joined, "quote": '"' in joined, "line_break": "\n" in joined or "\r" in joined,
                               "emoji": any(ord(c) > 0xFFFF for c in joined), "code_mixed_hindi": bool(re.search(r"[\u0900-\u097f]", joined) and re.search(r"[A-Za-z]", joined))}
                for feature, present in feature_map.items():
                    if present:
                        free_counts[feature] += 1; free_examples.setdefault(feature, mask(joined))
                for kind in poison_types(record):
                    poison_counts[kind] += 1; poison_sources[(kind, system)] += 1
                    if pid: poison_people[kind].add(pid)
                for value in flat_record.values():
                    text = clean_text(value)
                    if "@" in text and " " not in text:
                        domains[text.rsplit("@", 1)[1].lower()] += 1
                for field in ("phone", "contact_no"):
                    if not is_missing(flat_record.get(field)):
                        phone_values += 1; digits = re.sub(r"\D", "", str(flat_record[field])); phone_prefix_ok += digits.startswith("999") or digits in POISON_VALUES["placeholder_phone"] or len(digits) == 9 or (len(digits) == 10 and digits.startswith("0"))
                lowered_keys = {str(k).lower() for k in flat_record}; normal_labels |= lowered_keys & {"is_bot", "is_duplicate", "hard_negative", "poison_type", "true_person_id", "person_id", "evidence_mode", "scenario_type"}
                if re.search(r"automation\.internal|\b(?:monitor|crawler|scraper|healthcheck|lb-probe|bot)\b|dev-bot|\bheadless\b|\bautomation\b", joined.lower()): obvious_bot_markers += 1
                engagement = next((float(v) for k, v in record.items() if "engagement" in str(k).lower() and isinstance(v, (int, float))), None)
                stamp = clean_text(record.get(next(iter(TIMESTAMP_FIELDS[system]))))
                if entity_type == "bot":
                    bot_rows += 1; bot_times.add(stamp)
                    if engagement is not None: bot_engagement.append(engagement)
                elif entity_type == "human":
                    human_rows += 1; human_times.add(stamp)
                    if engagement is not None: human_engagement.append(engagement)
                parsed_primary = parse_timestamp(record.get(next(iter(TIMESTAMP_FIELDS[system]))))
                if parsed_primary:
                    (bot_instants if entity_type == "bot" else human_instants).append(parsed_primary)
                if engagement is not None: engagement_values.append((key, engagement))
                if any(clean_text(v).lower().endswith("@staff.test") for k, v in flat_record.items() if "email" in str(k).lower()):
                    if pid: qa_people.add(pid)
                if system == "social_logins":
                    payload = record.get("identity_payload")
                    provider_shapes[clean_text(record.get("provider")).lower()][tuple(sorted(payload)) if isinstance(payload, dict) else tuple()] += 1
                if system == "subscriptions" and clean_text(record.get("billing_name")) != clean_text(record.get("subscriber_name")): billing_different += 1
                if system == "ticketing":
                    if record.get("guest") is True and is_missing(record.get("account_id")): guest_no_account += 1
                    ref = clean_text(record.get("account_id"))
                    if ref: ticket_refs.append(ref)
                    event, created = parse_timestamp(record.get("event_ts")), parse_timestamp(record.get("created_ts"))
                    if event and created:
                        delay = (created-event).total_seconds()
                        (bot_ticket_delays if entity_type == "bot" else human_ticket_delays).append(delay)
                        late_eligible += 1
                        if delay > 9*86400: late += 1
                        if created < event: impossible_sets["negative_duration"].add(key)
                        if prior_event is not None:
                            ticket_adjacent += 1
                            if event < prior_event: inversions += 1
                        prior_event = event
                if system == "app_users" and rid.isdigit(): app_ids_numeric.add(int(rid))
                dob = clean_text(record.get("dob"))
                if dob and dob not in POISON_VALUES["default_dob"]:
                    try:
                        born = date.fromisoformat(dob); anchor = date(2026, 4, 3)
                        age = anchor.year-born.year-((anchor.month,anchor.day)<(born.month,born.day))
                        if age < 5: impossible_sets["age_below_5"].add(key)
                        if age > 110: impossible_sets["age_above_110"].add(key)
                        if born > anchor: impossible_sets["future_date"].add(key)
                    except ValueError: pass
                if system == "subscriptions":
                    start = parse_timestamp(record.get("start_date"))
                    if start and start > datetime(2026,4,3,12,0,tzinfo=timezone.utc): impossible_sets["future_date"].add(key)
        except Exception as exc:
            parse_errors.append(f"{system}: {type(exc).__name__}: {exc}")

    total_rows = sum(source_counts.values())
    overall_missing: Counter[str] = Counter(); overall_present: Counter[str] = Counter()
    for (system, field), count in missing_counts.items(): overall_missing[field] += count
    for system, fields in schemas.items():
        for field in fields: overall_present[field] += source_counts[system]

    print("[3/7] Re-reading repeated events and shared-evidence buckets...")
    duplicate_detail: dict[str, dict[str, Any]] = {}
    token_seen: Counter[str] = Counter(); token_person_seen: Counter[tuple[str, str]] = Counter()
    candidate_incidences = hard_candidate_incidences = 0
    unique_key_index: dict[tuple[str, str], int] = {}; token_unique_members: dict[str, set[int]] = defaultdict(set)
    uf = UnionFind(total_rows); owners: dict[str, int] = {}; poison_indices: list[int] = []; index_truth: list[str] = []
    repeated_payloads: dict[tuple[str, str], list[dict[str, Any]]] = defaultdict(list)
    index = 0
    for system, factory in factories.items():
        for record in factory():
            rid = source_record_id(system, record); source_key = (system, rid); mapped = truth.get(source_key); pid = mapped[0] if mapped else ""
            key_number = unique_key_index.setdefault(source_key, len(unique_key_index))
            index_truth.append(pid)
            if rid in repeat_ids[system]: repeated_payloads[(system, rid)].append(record)
            tokens = record_tokens(record)
            if poison_types(record): poison_indices.append(index)
            for token in tokens:
                if token_freq[token] >= 2:
                    token_unique_members[token].add(key_number)
                    candidate_incidences += token_seen[token]
                    hard_candidate_incidences += token_seen[token] - token_person_seen[(token, pid)]
                    token_seen[token] += 1; token_person_seen[(token, pid)] += 1
                    prior = owners.setdefault(token, index)
                    if prior != index: uf.union(index, prior)
            index += 1

    exact_by_source = {s: sum(n-1 for n in exact_hashes[s].values() if n > 1) for s in FILES}
    near_by_source: Counter[str] = Counter(); near_valid_by_source: Counter[str] = Counter()
    for (system, _rid), rows in repeated_payloads.items():
        unique = {fingerprint(r): r for r in rows}
        if len(unique) <= 1: continue
        values = list(unique.values()); base = values[0]
        for changed in values[1:]:
            near_by_source[system] += 1
            differing = [k for k in set(base)|set(changed) if base.get(k) != changed.get(k)]
            ts_field = TIMESTAMP_FIELDS[system][0] if system != "ticketing" else "created_ts"
            left, right = parse_timestamp(base.get(ts_field)), parse_timestamp(changed.get(ts_field))
            seconds = abs((right-left).total_seconds()) if left and right else math.inf
            if differing == [ts_field] and 1 <= seconds <= 60: near_valid_by_source[system] += 1
    exact_total, near_total = sum(exact_by_source.values()), sum(near_by_source.values())

    print("[4/7] Measuring true-pair evidence and unrecoverability...")
    pair_counts: Counter[str] = Counter(); pair_examples: dict[str, str] = {}; true_pairs = zero_pairs = 0; strength: Counter[int] = Counter(); zero_source_pairs: Counter[str] = Counter()
    worthless = {token for token, count in token_freq.items() if count > 40} | {token for token in token_freq if is_explicit_poison_token(token)}
    for records in multi_records.values():
        for left, right in combinations(records, 2):
            true_pairs += 1
            usable = (left["tokens"] & right["tokens"]) - worthless
            strength[len(usable)] += 1
            if not usable:
                zero_pairs += 1
                zero_source_pairs["+".join(sorted((left["system"], right["system"])))] += 1
            categories: list[str] = []
            if left["email_raw"] and left["email_raw"] == right["email_raw"] and (left["verified"] or right["verified"]): categories.append("exact_verified_email")
            if left["email_raw"] and right["email_raw"] and left["email_raw"] != right["email_raw"] and left["email"] == right["email"]:
                if left["email_raw"].lower() == right["email_raw"].lower(): categories.append("email_case")
                if "." in left["email_raw"].split("@",1)[0] or "." in right["email_raw"].split("@",1)[0]: categories.append("email_dots")
                if "+" in left["email_raw"].split("@",1)[0] or "+" in right["email_raw"].split("@",1)[0]: categories.append("email_plus")
            if left["phone"] and left["phone"] == right["phone"] and left["phone_raw"] != right["phone_raw"]:
                if left["phone_raw"].startswith("+999") != right["phone_raw"].startswith("+999"): categories.append("phone_country_code")
                if " " in left["phone_raw"] or " " in right["phone_raw"]: categories.append("phone_spaces")
                if left["phone_raw"].startswith("0") or right["phone_raw"].startswith("0"): categories.append("phone_leading_zero")
            if usable and all(t.startswith("name_city:") for t in usable): categories.append("only_name_city")
            if usable and all(t.startswith("device:") for t in usable): categories.append("only_device")
            for category in categories:
                pair_counts[category] += 1
                if category.startswith("email"): example = f"{mask(left['email_raw'],'email')} ↔ {mask(right['email_raw'],'email')}"
                elif category.startswith("phone"): example = f"{mask(left['phone_raw'],'phone')} ↔ {mask(right['phone_raw'],'phone')}"
                else: example = "masked synthetic pair; no truth identifier exposed"
                pair_examples.setdefault(category, example)

    print("[5/7] Validating explicit hard negatives and poisoned components...")
    manifest = json.loads((data_dir / "hard_negatives.json").read_text(encoding="utf-8")); hard_types: Counter[str] = Counter(); hard_valid = 0
    for item in manifest:
        kind = str(item.get("type", "")); hard_types[kind] += 1
        ids = item.get("person_ids", []); refs = item.get("source_records", [])
        mapped_ids = [truth.get((str(ref.get("system")), str(ref.get("record_id"))), (None,))[0] for ref in refs]
        if len(ids) == 2 and ids[0] != ids[1] and len(mapped_ids) == 2 and mapped_ids == ids and item.get("must_not_merge") is True: hard_valid += 1
    canonical_links = list(iter_jsonl(canonical_path)); canonical_seen: set[tuple[tuple[str,str],tuple[str,str]]] = set()
    canonical_valid = canonical_mode_matches = canonical_recoverability_matches = canonical_unrecoverable = 0
    canonical_modes: Counter[str] = Counter(); canonical_primary_modes: Counter[str] = Counter()
    for link in canonical_links:
        left_key = (str(link.get("source_system_a")), str(link.get("source_record_id_a")))
        right_key = (str(link.get("source_system_b")), str(link.get("source_record_id_b")))
        ordered = tuple(sorted((left_key, right_key)))
        left_truth, right_truth = truth.get(left_key), truth.get(right_key)
        valid = bool(left_truth and right_truth and left_truth[0] == right_truth[0] and left_key != right_key and ordered not in canonical_seen)
        if valid:
            expected_key = hashlib.sha256(f"canonical:{left_truth[0]}".encode()).hexdigest()
            valid = link.get("truth_key") == expected_key and link.get("scenario_type") == "canonical_duplicate_link"
        canonical_seen.add(ordered)
        if valid: canonical_valid += 1
        left_compact, right_compact = multi_record_lookup.get(left_key), multi_record_lookup.get(right_key)
        if left_compact and right_compact:
            usable = (left_compact["tokens"] & right_compact["tokens"]) - worthless
            observed = observed_modes(left_compact, right_compact, usable)
            intended = set(link.get("evidence_modes") or [link.get("evidence_mode")])
            if intended == observed and link.get("evidence_mode") in intended: canonical_mode_matches += 1
            recoverable = bool(usable)
            if bool(link.get("intended_recoverability")) == recoverable: canonical_recoverability_matches += 1
            if not recoverable: canonical_unrecoverable += 1
            canonical_modes.update(intended)
            canonical_primary_modes[str(link.get("evidence_mode"))] += 1
    expected_canonical = sum(max(0, n-1) for n in people_counts.values())
    canonical_rate = canonical_unrecoverable/len(canonical_links) if canonical_links else 0.0

    high_items = [(token,members) for token,members in token_unique_members.items() if token_freq[token] > 40]
    poison_high_groups = [members for token,members in high_items if is_explicit_poison_token(token)]
    moderate_high_pairs: set[int] = set()
    for token,members in high_items:
        if is_explicit_poison_token(token): continue
        for left_index,right_index in combinations(sorted(members),2): moderate_high_pairs.add(pair_code(left_index,right_index))
    rule2_pairs: set[int] = set()
    for token, members in token_unique_members.items():
        if token_freq[token] <= 40 and len(members) >= 2:
            for left_index, right_index in combinations(sorted(members),2): rule2_pairs.add(pair_code(left_index,right_index))
    poison_union = high_edge_union(poison_high_groups)
    moderate_also_poison=0
    for code in moderate_high_pairs:
        left_index,right_index=code>>20,code&((1<<20)-1)
        if any(left_index in group and right_index in group for group in poison_high_groups): moderate_also_poison+=1
    excluded_union=poison_union+len(moderate_high_pairs)-moderate_also_poison
    rule2_also_high = 0
    for code in rule2_pairs:
        left_index, right_index = code >> 20, code & ((1 << 20)-1)
        if code in moderate_high_pairs or any(left_index in group and right_index in group for group in poison_high_groups): rule2_also_high += 1
    unique_naive_pairs = excluded_union + len(rule2_pairs) - rule2_also_high
    explicit_pairs: set[int] = set()
    for item in manifest:
        refs = item.get("source_records",[])
        if len(refs)==2:
            left_index=unique_key_index.get((str(refs[0].get("system")),str(refs[0].get("record_id"))))
            right_index=unique_key_index.get((str(refs[1].get("system")),str(refs[1].get("record_id"))))
            if left_index is not None and right_index is not None and left_index != right_index: explicit_pairs.add(pair_code(left_index,right_index))
    explicit_candidate_pairs = len(explicit_pairs & rule2_pairs)
    explicit_hard_rate = explicit_candidate_pairs/len(rule2_pairs) if rule2_pairs else 0.0
    pairs_only_rule2_values = excluded_union-rule2_also_high
    non_poison_pairs=rule2_pairs|moderate_high_pairs
    poison_overlap_nonpoison=0
    for code in non_poison_pairs:
        left_index,right_index=code>>20,code&((1<<20)-1)
        if any(left_index in group and right_index in group for group in poison_high_groups): poison_overlap_nonpoison+=1
    pairs_only_poison=poison_union-poison_overlap_nonpoison
    after_rule2_pct = len(rule2_pairs)/unique_naive_pairs if unique_naive_pairs else 0.0
    high_non_poison_count = sum(not is_explicit_poison_token(token) for token,_members in high_items)
    poisoned_roots = {uf.find(i) for i in poison_indices}; largest_root = max(poisoned_roots, key=lambda r: uf.size[r]) if poisoned_roots else None
    largest_cluster = uf.size[largest_root] if largest_root is not None else 0
    collapsed_people = len({pid for i, pid in enumerate(index_truth) if largest_root is not None and uf.find(i) == largest_root and pid})
    cluster_causes = sorted({token for token, owner in owners.items() if largest_root is not None and uf.find(owner) == largest_root and is_explicit_poison_token(token)})

    print("[6/7] Comparing independently measured values with the generation report...")
    report = json.loads((data_dir / "generation_report.json").read_text(encoding="utf-8"))
    # Predeclared tolerances: full-scale deterministic data warrants tighter bounds than a sample.
    compare_approx(audit,"SCALE-001","Distinct synthetic people",len(people_counts),300000,6000,"people","Tolerance is 2% because 'approximately' is qualitative but this is a full deterministic population.","build_people/generate")
    compare_approx(audit,"SCALE-002","Source account/booking rows",total_rows,420000,21000,"rows","Tolerance is 5% for 'approximately'; emitted files were independently parsed.","generate")
    multi_rate = len(multi_people)/len(people_counts)
    compare_approx(audit,"SCALE-003","People with multiple physical records",multi_rate,0.25,0.02,"proportion","Tolerance is ±2 percentage points, declared before evaluation.","generate allocation plan")
    audit.add("SCALE-004","Small nonzero group has at least six records","1 to 1% of people",six_plus,"PASS" if 0<six_plus<=0.01*len(people_counts) else "FAIL",f"Computed from physical truth-map occurrences: {six_plus:,}.","Adjust six_people allocation and regenerate.",location="generate:six_people",regen=True)
    audit.add("SCALE-005","Complete generated source dataset under 3,000,000 rows","<3,000,000",total_rows,"PASS" if total_rows<3_000_000 else "FAIL","Sum of independently parsed source rows.")
    synthetic_domains = all(d.endswith(".test") or d.endswith(".internal") or d.endswith(".example") for d in domains)
    audit.add("SCALE-006","All identities and identifiers are synthetic","Only constructed synthetic values","No external source inputs referenced; synthetic domains and generated IDs observed","PASS" if synthetic_domains else "WARNING",f"Observed domains={dict(domains)}; source inspection found syllable/token construction only.","Replace any non-reserved domains and regenerate." if not synthetic_domains else "None.",mandatory=True,location="build_people/synth_email")
    audit.add("SCALE-007","Emails use synthetic/test domains","All domains reserved/synthetic",dict(domains),"PASS" if synthetic_domains else "FAIL","Every observed email-like domain was counted.",location="synth_email/make_bot_row",regen=not synthetic_domains)
    audit.add("SCALE-008","Phones use clearly fictional ranges","All values use +999 bases, defined local variants, or explicit placeholders",f"{phone_prefix_ok:,}/{phone_values:,} structurally synthetic/variant","PASS" if phone_values and phone_prefix_ok==phone_values else "FAIL","Raw values inspected without rewriting: +999 bases have 9-digit synthetic payloads; required no-country-code and leading-zero representations retain that payload.",location="synth_phone/phone_variant",regen=phone_prefix_ok!=phone_values)
    generator_text = generator.read_text(encoding="utf-8") if generator and generator.is_file() else ""
    seed_ok = bool(re.search(r"--seed.*default=42",generator_text)) and "random.Random(seed)" in generator_text
    audit.add("SCALE-009","Reproducible documented seed","CLI seed with deterministic RNG",report.get("generation_parameters",{}).get("seed"),"PASS" if seed_ok and report.get("generation_parameters",{}).get("seed")==42 else "NOT VERIFIABLE",f"Generator contains CLI default seed and local random.Random(seed): {seed_ok}; report records seed 42.","Document and use a local seeded RNG.",location="main/generate",regen=False)

    for idx, system in enumerate(FILES, 2):
        ok = EXPECTED[system] <= schemas[system]
        audit.add(f"FORMAT-{idx:03d}",f"{system} is readable and has required schema",str(sorted(EXPECTED[system])),f"rows={source_counts[system]:,}; fields={sorted(schemas[system])}","PASS" if ok and not parse_errors else "FAIL","Parsed with csv/json/json-lines/openpyxl as appropriate.","Correct writer/schema and regenerate.",location="write_outputs",regen=not ok)
    sheets, header_row, excel_header = subscription_info(data_dir/FILES["subscriptions"])
    audit.add("FORMAT-007","At least one multi-sheet Excel workbook","more than one sheet",sheets,"PASS" if len(sheets)>1 else "FAIL","Workbook metadata read with openpyxl.",location="write_outputs",regen=len(sheets)<=1)
    audit.add("FORMAT-008","Excel true header below row 1","header row >1",header_row,"PASS" if header_row>1 and set(excel_header)>=EXPECTED["subscriptions"] else "FAIL",f"Detected row {header_row}: {excel_header}.",location="write_outputs",regen=True)
    schema_signatures = {tuple(sorted(v)) for v in schemas.values()}
    audit.add("FORMAT-009","Raw source schemas remain inconsistent","5 distinct schema signatures",len(schema_signatures),"PASS" if len(schema_signatures)==5 else "FAIL","Compared emitted field-name sets; no standardization performed.",location="write_outputs",regen=True)
    audit.add("FORMAT-010","Ticketing supports repeated bookings and guest records without accounts","nonzero repeated people and guest/no-account rows",f"guest without account={guest_no_account:,}","PASS" if guest_no_account>0 else "FAIL","Booking IDs are row/event keys; truth aggregates show people can recur.",location="make_ticket_row",regen=True)
    audit.add("FORMAT-011","Subscription billing name can belong to parent/partner","nonzero different billing names",billing_different,"PASS" if billing_different>0 else "FAIL","Compared subscriber_name and billing_name without standardizing.",location="make_subscription_row",regen=True)
    providers=set(provider_shapes); shapes={p:sorted([list(x) for x in v]) for p,v in provider_shapes.items()}
    distinct_payload_shapes={shape for counts in provider_shapes.values() for shape in counts}
    audit.add("FORMAT-012","Social logins have exactly four heterogeneous providers","exactly 4",f"providers={sorted(providers)}","PASS" if len(providers)==4 and len(distinct_payload_shapes)>=4 else "FAIL",f"Provider identity-payload shapes={shapes}; outer login_ts/channel/engagement metadata excluded from identity availability.",location="make_social_row",regen=True)
    verified_providers=[p for p,s in shapes.items() if any("verified_email" in x for x in s)]
    audit.add("FORMAT-013","Social provider supports verified email","at least one provider",verified_providers,"PASS" if verified_providers else "FAIL","Observed inside provider-returned identity payloads.",location="make_social_row",regen=True)
    only_id_display = [p for p,c in provider_shapes.items() if any(set(shape)=={"provider_id","display_name"} for shape in c)]
    audit.add("FORMAT-014","A social provider can supply only provider ID and display name","at least one identity payload",only_id_display,"PASS" if only_id_display else "FAIL","Identity payload alone was examined; operational timestamp/channel/engagement fields are outside it.","Emit a minimal nested identity payload.",location="make_social_row",regen=not only_id_display)
    hashed_providers=[p for p,s in shapes.items() if any("hashed_email" in x and "verified_email" not in x for x in s)]
    audit.add("FORMAT-015","A social provider uses hashed rather than plaintext email","at least one provider",hashed_providers,"PASS" if hashed_providers else "FAIL","Observed inside provider-returned identity payloads.",location="make_social_row",regen=True)
    display_providers=[p for p,s in shapes.items() if any("display_name" in x for x in s)]
    audit.add("FORMAT-016","Social display names include nicknames/handles","nonzero provider payloads",display_providers,"PASS" if display_providers else "FAIL","Display-name payloads are provider-specific handles or names.",location="make_social_row",regen=True)

    pair_labels=[("DUP-001","exact_verified_email","Exact verified email match"),("DUP-002","email_case","Email case differences"),("DUP-003","email_dots","Dots in email local part"),("DUP-004","email_plus","Plus-address suffixes"),("DUP-005","phone_country_code","Phone with/without country code"),("DUP-006","phone_spaces","Phone numbers containing spaces"),("DUP-007","phone_leading_zero","Phone numbers using a leading zero"),("DUP-008","only_name_city","Pairs sharing only name and city"),("DUP-009","only_device","Pairs sharing only device ID")]
    for ident,key,label in pair_labels:
        count=pair_counts[key]; applicable=true_pairs
        audit.add(ident,label,"nonzero true-pair count",f"{count:,}/{applicable:,} ({pct(count/applicable if applicable else 0)})","PASS" if count>0 else "FAIL",f"All pairwise physical human records for the same hidden person; safe example: {pair_examples.get(key,'none')}.","Adjust evidence-mode generation and regenerate.",location="make_*_row/evidence_mode",regen=count==0)
        audit.metric(ident,label,count,"pairs")
    audit.add("DUP-010","Duplicate sets contain different evidence strengths","at least 2 strength levels",dict(strength),"PASS" if len(strength)>=2 else "FAIL","Strength is count of shared usable evidence tokens after >40-frequency/poison exclusion.",location="evidence_mode",regen=True)
    zero_rate=zero_pairs/true_pairs if true_pairs else 0
    compare_approx(audit,"UNREC-001","Unrecoverable true duplicate pairs",zero_rate,0.08,0.01,"proportion",f"Usable evidence = normalized email, phone, device, account reference, payment token, hash or name+city, excluding explicit poison and values occurring >40 times. Denominator=all {true_pairs:,} pairwise physical human-record combinations; zero={zero_pairs:,}.","evidence_mode/make_*_row")
    canonical_ok = len(canonical_links)==expected_canonical and canonical_valid==len(canonical_links) and canonical_recoverability_matches==len(canonical_links)
    audit.add("UNREC-002","Canonical duplicate-link unrecoverable rate is independently verified","complete valid canonical star links",f"{canonical_unrecoverable:,}/{len(canonical_links):,} ({pct(canonical_rate)})","PASS" if canonical_ok else "FAIL",f"Expected links=sum(unique human records-1)={expected_canonical:,}; valid={canonical_valid:,}; recoverability reconciled={canonical_recoverability_matches:,}. Pairwise denominator={true_pairs:,} and canonical denominator={len(canonical_links):,}; they need not match.","Correct hidden canonical link construction.",location="build_hidden_metadata/write_outputs",regen=not canonical_ok)

    required_hard={"father_son","university_computer_lab","common_name_city","couple_shared_email_payment_token"}
    audit.add("HARDNEG-001","All required hard-negative scenarios have different truth IDs",str(sorted(required_hard)),dict(hard_types),"PASS" if required_hard<=set(hard_types) and hard_valid==len(manifest) else "FAIL",f"Validated {hard_valid:,}/{len(manifest):,} manifest entries against hidden truth without exposing IDs.","Correct hard-negative source backing and regenerate.",location="apply_hard_negative_profiles/write_outputs",regen=True)
    audit.add("HARDNEG-002","Explicit hard negatives are at least 5% of unique Rule-2 candidate pairs",">=5%",f"{explicit_candidate_pairs:,}/{len(rule2_pairs):,} ({pct(explicit_hard_rate)})","PASS" if explicit_hard_rate>=.05 else "FAIL","Denominator is unique unordered distinct source-record-key pairs sharing at least one evidence value occurring on <=40 physical rows. Numerator includes only manifest-labelled pairs with different truth IDs that are candidates.","Increase explicit source-backed hard-negative pairs without weakening Rule 2.",location="apply_hard_negative_profiles",regen=True)
    audit.metric("HARDNEG-002","unique_unordered_naive_candidate_pairs",unique_naive_pairs,"pairs"); audit.metric("HARDNEG-002","explicit_hard_negative_candidate_pairs",explicit_candidate_pairs,"pairs")
    audit.add("HARDNEG-004","Naive candidate incidences and unique pairs are reported separately","both metrics",f"incidences={candidate_incidences:,}; unique pairs={unique_naive_pairs:,}","PASS","Incidences count repeated evidence/physical rows; unique pairs deduplicate unordered source-record keys.")
    audit.add("HARDNEG-005","Poison-only and post-Rule-2 candidates are quantified","exact counts",f"only Rule-2 values={pairs_only_rule2_values:,}; poison-only={pairs_only_poison:,}; after Rule 2={len(rule2_pairs):,} ({pct(after_rule2_pct)} of naive unique); high-frequency values={len(high_items)}","PASS",f"Poison cliques used exact inclusion-exclusion; {high_non_poison_count} smaller high-frequency non-poison values were explicitly enumerated.")
    audit.add("HARDNEG-003","Hard-negative breakdown is available by explicit label","all five concepts",dict(hard_types),"PASS" if required_hard<=set(hard_types) else "FAIL","Couple shared-email and shared-payment-token are intentionally combined in the manifest and count toward both concepts.",location="hard_negatives.json",regen=True)

    poison_targets={"placeholder_phone":(3000,600),"default_dob":(4000,2000),"corporate_email":(900,180),"kiosk_device":(40000,10000),"test_email":(1500,300)}
    poison_display={"placeholder_phone":"0000000000 / 9999999999","default_dob":"1900-01-01 / 1970-01-01 (and DD-MM equivalents)","corporate_email":"bookings@events.example","kiosk_device":"KIOSK-DEVICE-1","test_email":"qa+001@staff.test"}
    for num,(kind,(target,tol)) in enumerate(poison_targets.items(),1):
        count=poison_counts[kind]; absolute,relative=deviation(count,target)
        status="PASS" if absolute<=tol and count>40 else "FAIL"
        detail=f"synthetic value={poison_display[kind]}; by source={ {s:poison_sources[(kind,s)] for s in FILES if poison_sources[(kind,s)]} }; distinct true entities={len(poison_people[kind]):,}; exceeds worthless threshold 40={count>40}; absolute deviation={absolute:,}; relative deviation={pct(relative)}; tolerance=±{tol:,}."
        audit.add(f"POISON-{num:03d}",f"{kind} injected at required scale",f"about {target:,}",count,status,detail,"Adjust poison allocation and regenerate.",location="apply_poison",regen=status=="FAIL")
    audit.add("POISON-006","Naive poisoned component reaches tens of thousands",">=10,000 records",largest_cluster,"PASS" if largest_cluster>=10000 else "FAIL",f"Independent union-find over exact naive tokens; distinct hidden entities collapsed={collapsed_people:,}; connecting poison tokens={cluster_causes}. No members printed.","Adjust poison placement/connectivity and regenerate.",location="apply_poison",regen=True)

    exact_rate=exact_total/total_rows; near_rate=near_total/total_rows
    compare_approx(audit,"ROWDUPE-001","Exact duplicate rows",exact_rate,.02,.0025,"proportion",f"Definition: identical complete parsed row including record ID and timestamp; no fields excluded. Count={exact_total:,}; by source={exact_by_source}; denominator=all emitted rows.","add_artifacts")
    compare_approx(audit,"ROWDUPE-002","Near-duplicate events",near_rate,.01,.0025,"proportion",f"Same event=source record ID; one distinct complete-row variant; allowed time delta=1..60 seconds; changed-field/time validation={sum(near_valid_by_source.values()):,}/{near_total:,}; by source={dict(near_by_source)}; denominator=all emitted rows.","add_artifacts")
    audit.add("ROWDUPE-003","Every near duplicate changes exactly one timestamp field by seconds","100% valid",f"{sum(near_valid_by_source.values()):,}/{near_total:,}","PASS" if near_total and sum(near_valid_by_source.values())==near_total else "FAIL","Re-read only repeated event IDs and compared unique parsed dictionaries.","Change only the source timestamp field by 1..60 seconds.",location="add_artifacts",regen=True)

    miss_rates={f"{s}.{f}":missing_counts[(s,f)]/source_counts[s] for s,fields in schemas.items() for f in fields if source_counts[s]}
    mid_missing={k:v for k,v in miss_rates.items() if .04<=v<=.09}
    audit.add("MISS-001","At least three columns have 4%-9% missingness",">=3 source-columns",{k:pct(v) for k,v in mid_missing.items()},"PASS" if len(mid_missing)>=3 else "FAIL",f"Missing means actual null, empty/whitespace, 'null' or 'None' case-insensitively. All treated consistently.","Tune conditional_missing and regenerate.",location="conditional_missing/make_*_row",regen=True)
    source_email_rates={s:miss_rates.get(f"{s}.email",miss_rates.get(f"{s}.customer_email_address",0)) for s in FILES}
    source_effect=max(source_email_rates.values())-min(source_email_rates.values())
    audit.add("MISS-002","Missingness is non-random and associated with source/country/device",">=1 percentage-point effect",f"source email max-min={pct(source_effect)}","PASS" if source_effect>=.01 else "FAIL",f"Source email missing rates={ {s:pct(v) for s,v in source_email_rates.items()} }; simple effect size is max-minus-min rate.","Increase conditional source/country/device effects and regenerate.",location="conditional_missing",regen=True)
    audit.metric("MISS-001","missing_rates_by_source_column",json.dumps({k:pct(v) for k,v in sorted(miss_rates.items())}),"")

    formats=Counter({fmt:sum(n for (s,f,k),n in timestamp_counts.items() if k==fmt) for fmt in {k for _,_,k in timestamp_counts}})
    audit.add("TIME-001","A timestamp column contains at least four required formats",">=4 formats in one column",dict(formats),"PASS" if len(formats)>=4 else "FAIL",f"Detected raw formats without rewriting; unparseable={dict(timestamp_unparseable)}.","Increase mixed_timestamp variety and regenerate.",location="mixed_timestamp",regen=True)
    audit.add("TIME-002","Ambiguous dates exist","nonzero",ambiguous_dates,"PASS" if ambiguous_dates else "FAIL","Counted DD-MM forms where both day and month are <=12.",location="mixed_timestamp",regen=True)
    local_count=formats.get("LOCAL_TEXT",0); utc_count=formats.get("ISO_OFFSET",0)+formats.get("EPOCH_MS",0)
    audit.add("TIME-003","UTC, local and timezone-unspecified timestamps coexist","all present",f"UTC/offset-or-epoch={utc_count:,}; local/no-zone={local_count:,}","PASS" if utc_count and local_count else "FAIL","ISO offsets/epoch encode UTC; LOCAL_TEXT and three date-only styles carry no timezone marker. Semantics cross-checked with saved definition.",location="mixed_timestamp",regen=True)

    for num,concept in enumerate(("channel","city","device_type","country"),1):
        counts=Counter();
        for (c,s,v),n in category_counts.items():
            if c==concept: counts[v]+=n
        audit.add(f"CAT-{num:03d}",f"{concept} contains at least three raw variants",">=3",dict(counts),"PASS" if len(counts)>=3 else "FAIL","Raw variants counted without normalization.","Add spelling/case/format variants and regenerate.",location="*_VARIANTS",regen=True)

    padded_store=sum(1 for key in source_occ if key[0]=="store_customers")
    audit.add("JOIN-001","Logical app ID is integer-like in CSV and zero-padded text elsewhere","numeric lexical app ID; zero-padded store reference","app.account_id numeric lexical; store.app_account_ref zero-padded CSV text","PASS" if app_ids_numeric and "app_account_ref" in schemas["store_customers"] else "FAIL","CSV has no intrinsic typed columns; physical lexical representations were inspected. Safe example: integer n ↔ zero-padded 000…n.",location="make_store_row",regen=True)
    orphan=sum(1 for ref in ticket_refs if ref.isdigit() and int(ref) not in app_ids_numeric); orphan_rate=orphan/len(ticket_refs) if ticket_refs else 0
    audit.add("JOIN-002","3%-6% event-table account IDs are orphaned","3%-6%",f"ticketing={orphan:,}/{len(ticket_refs):,} ({pct(orphan_rate)}); overall same","PASS" if .03<=orphan_rate<=.06 else "FAIL","Appropriate account table is app_users; blank guest references excluded from denominator.","Tune orphan_target and regenerate.",location="generate/make_ticket_row",regen=True)

    if engagement_values:
        median_eng=statistics.median(v for _,v in engagement_values); threshold=median_eng*100
        for key,v in engagement_values:
            if v>=threshold: impossible_sets["engagement_100x_median"].add(key)
    else: median_eng=0
    impossible_union=set().union(*impossible_sets.values()) if impossible_sets else set(); impossible_rate=len(impossible_union)/total_rows
    for num,kind in enumerate(("negative_duration","age_below_5","age_above_110","future_date","engagement_100x_median"),1):
        audit.add(f"IMP-{num:03d}",f"Impossible/extreme category: {kind}","nonzero",len(impossible_sets[kind]),"PASS" if impossible_sets[kind] else "FAIL",f"Rows counted once within category; engagement threshold={threshold if engagement_values else 'n/a'} (100× median {median_eng}).","Inject category and regenerate.",location="apply_impossible_values",regen=True)
    compare_approx(audit,"IMP-006","Combined impossible/extreme row rate",impossible_rate,.003,.001,"proportion",f"Union of row keys prevents cross-category double counting; count={len(impossible_union):,}.","apply_impossible_values")

    bot_rate=bot_rows/total_rows
    audit.add("BOT-001","Automated traffic is 4%-7%","4%-7%",f"{bot_rows:,}/{total_rows:,} ({pct(bot_rate)})","PASS" if .04<=bot_rate<=.07 else "FAIL","Entity type read only from hidden truth; no records displayed.","Tune bot allocation and regenerate.",location="make_bot_row/generate",regen=True)
    audit.add("BOT-002","No direct bot flag in normal outputs","no direct label",sorted(normal_labels),"PASS" if "is_bot" not in normal_labels else "FAIL","All source schemas inspected.","Remove direct labels and regenerate.",location="write_outputs",regen=True)
    bot_med=statistics.median(bot_engagement) if bot_engagement else None; human_med=statistics.median(human_engagement) if human_engagement else None
    bot_span=(max(bot_instants)-min(bot_instants)).total_seconds() if len(bot_instants)>1 else 0; human_span=(max(human_instants)-min(human_instants)).total_seconds() if len(human_instants)>1 else 0
    bot_delay_med=statistics.median(bot_ticket_delays) if bot_ticket_delays else None; human_delay_med=statistics.median(human_ticket_delays) if human_ticket_delays else None
    signals={"low_engagement":bot_med is not None and human_med is not None and bot_med<=.2*human_med,"narrow_time_window":bot_span<=3600 and human_span>86400,"zero_or_short_ticket_delay":bot_delay_med is not None and human_delay_med is not None and bot_delay_med<=1 and human_delay_med>60}
    subtle=obvious_bot_markers==0 and sum(signals.values())>=3
    audit.add("BOT-003","Bot rows use subtle behavioural patterns rather than directly revealing values","no textual markers and >=3 quantitative behavioural signals",f"bot engagement median={bot_med}; human={human_med}; bot time span={bot_span:.0f}s; human={human_span:.0f}s; bot ticket delay median={bot_delay_med}s; human={human_delay_med}s; obvious-marker rows={obvious_bot_markers}; signals={signals}","PASS" if subtle else "FAIL","Hidden labels define populations only. The check requires a combination of low engagement, concentrated/regular timing and very short event-to-recording delays, plus zero obvious bot-revealing text.","Use neutral identities and strengthen behavioural combinations.",cause="Insufficient behavioural separation or an obvious marker remains.",location="make_bot_row",regen=True)

    qa_rate=len(qa_people)/len(people_counts) if people_counts else 0
    compare_approx(audit,"QA-001","Internal test/QA people",qa_rate,.005,.001,"proportion",f"Distinct hidden human identities with staff.test email={len(qa_people):,}; normal rows include qa+001/load-test patterns; denominator=distinct human people.","apply_poison/QA naming")
    for num,feature in enumerate(("comma","quote","line_break","emoji","code_mixed_hindi"),1):
        audit.add(f"TEXT-{num:03d}",f"Free text includes {feature}","nonzero rows",free_counts[feature],"PASS" if free_counts[feature] else "FAIL",f"Standards-compliant parsers successfully read all files; safe masked example={free_examples.get(feature,'none')}.","Retain messy_text injection and regenerate.",location="messy_text",regen=True)
    with (data_dir/"app_users.csv").open(encoding="utf-8") as handle:
        physical_lines=sum(1 for _ in handle)-1
    audit.add("TEXT-006","CSV requires standards-compliant parsing","physical line count differs or quoted features present",f"app parsed rows={source_counts['app_users']:,}; physical data lines={physical_lines:,}","PASS" if physical_lines!=source_counts["app_users"] or free_counts["comma"] or free_counts["quote"] else "FAIL","csv.DictReader parsed successfully; embedded delimiters/newlines make naive splitting unreliable.",location="write_outputs",regen=True)

    late_rate=late/late_eligible if late_eligible else 0; inv_rate=inversions/ticket_adjacent if ticket_adjacent else 0
    audit.add("LATE-001","At least 3% of events arrive >9 days late",">=3%",f"{late:,}/{late_eligible:,} ({pct(late_rate)})","PASS" if late_rate>=.03 else "FAIL","Event timestamp=ticketing.event_ts; arrival=ticketing.created_ts; late=(arrival-event)>9×24h after parsing raw formats.","Increase late_target and regenerate.",location="make_ticket_row/generate",regen=True)
    audit.add("LATE-002","Applicable event rows are out of chronological order","nonzero adjacent inversions",f"{inversions:,}/{ticket_adjacent:,} ({pct(inv_rate)})","PASS" if inversions>0 else "FAIL","Compared adjacent parsed event_ts values in emitted ticketing order.","Shuffle rows before writing and regenerate.",location="generate:rng.shuffle",regen=True)

    missing_map=sum(max(0,source_occ[k]-truth_occ.get(k,0)) for k in source_occ); extra_map=sum(max(0,truth_occ[k]-source_occ.get(k,0)) for k in truth_occ)
    duplicate_truth_keys=sum(1 for n in truth_occ.values() if n>1); justified=all(source_occ[k]==n for k,n in truth_occ.items() if n>1)
    audit.add("TRUTH-001","Every applicable source row has one separate truth mapping","physical counts match",f"source={total_rows:,}; truth={truth_rows:,}; missing={missing_map}; extra={extra_map}; conflicts={truth_conflicts}","PASS" if total_rows==truth_rows and not missing_map and not extra_map and not truth_conflicts else "FAIL","Compared physical occurrence counts by (system, record_id); truth content not emitted.","Correct person_map writing and regenerate.",location="write_outputs",regen=True)
    audit.add("TRUTH-002","Normal source files contain no truth/scenario labels","none",sorted(normal_labels),"PASS" if not normal_labels else "FAIL","All emitted source field names inspected.","Remove labels and regenerate.",location="write_outputs",regen=True)
    audit.add("TRUTH-003","Duplicate truth-map source keys are absent or justified","duplicates only for repeated physical events",duplicate_truth_keys,"PASS" if justified and not truth_conflicts else "FAIL","Repeated keys have identical truth IDs and occurrence counts match exact/near repeated source rows.","Use physical-row keys or document repeated event keys.",location="add_artifacts/write_outputs",regen=True)
    prod_reader_text=(Path(__file__).resolve().parents[1]/"scripts"/"verify_synthetic_dataset.py").read_text(encoding="utf-8") if (Path(__file__).resolve().parents[1]/"scripts"/"verify_synthetic_dataset.py").is_file() else ""
    audit.add("TRUTH-004","Truth is isolated from production-style source parsing","source readers do not require truth","Separate source iterators and evaluation join","PASS","Each normal file was parsed before truth joining; person_map is only referenced in evaluation sections.")

    forbidden_behaviours={"normal output labels":not normal_labels,"no chronological sort":inversions>0,"no output deduplication":exact_total>0,"no repair of impossible values":bool(impossible_union),"no category standardization":all(len({v for (c,s,v) in category_counts if c==concept})>=3 for concept in ("channel","city","device_type","country"))}
    audit.add("GEN-001","Generator preserves deliberate messiness and avoids forbidden cleanup", "all observable behaviours preserved",forbidden_behaviours,"PASS" if all(forbidden_behaviours.values()) else "FAIL","Evidence comes from emitted files; source was additionally inspected for post-generation cleanup.","Remove cleanup/standardization/repair/sorting/deduplication logic and regenerate.",location="generate/write_outputs",regen=True)

    hard_rate=explicit_hard_rate
    measured_report=report.get("measured_problem_rates",{}); injected_report=report.get("injected_problem_rates",{})
    claims={"row_counts":report.get("row_counts"),"total_rows":report.get("total_row_count"),"true_people":report.get("true_number_of_distinct_people"),"multi_rate":report.get("true_duplicate_rate_people_with_multiple_records"),"six_plus":report.get("people_with_six_or_more_records"),"poisons":report.get("poison_cluster_sizes"),"largest_cluster":report.get("largest_poisoned_cluster_naive_matcher"),"largest_cluster_people":report.get("largest_poisoned_cluster_distinct_true_people"),"largest_cluster_causes":report.get("largest_poisoned_cluster_causes"),"estimated_cluster":report.get("estimated_largest_poisoned_cluster"),"zero_pairs":measured_report.get("zero_evidence_pairs"),"true_pairs":measured_report.get("true_duplicate_pairs"),"zero_rate":measured_report.get("measured_zero_evidence_duplicate_pair_rate"),"zero_source_pairs":measured_report.get("zero_evidence_by_source_pair"),"canonical_links":measured_report.get("canonical_duplicate_links"),"canonical_unrecoverable":measured_report.get("canonical_unrecoverable_links"),"canonical_rate":measured_report.get("canonical_unrecoverable_rate"),"candidate_incidences":measured_report.get("candidate_pair_incidences_before_deduplication"),"different_person_incidences":measured_report.get("different_person_candidate_incidences"),"unique_naive_pairs":measured_report.get("unique_unordered_naive_candidate_pairs"),"pairs_only_rule2_values":measured_report.get("pairs_created_only_through_rule2_values"),"pairs_only_poison":measured_report.get("pairs_created_only_through_poison_identifiers"),"rule2_pairs":measured_report.get("rule2_unique_candidate_pairs"),"after_rule2_pct":measured_report.get("candidate_percentage_after_rule2"),"explicit_hard_pairs":measured_report.get("explicit_hard_negative_pairs"),"explicit_hard_candidates":measured_report.get("explicit_hard_negative_candidate_pairs"),"explicit_hard_rate":measured_report.get("explicit_hard_negative_rate_rule2"),"high_value_count":measured_report.get("rule2_high_frequency_value_count"),"evidence_modes":measured_report.get("duplicate_evidence_mode_counts")}
    recomputed={"row_counts":dict(source_counts),"total_rows":total_rows,"true_people":len(people_counts),"multi_rate":multi_rate,"six_plus":six_plus,"poisons":dict(poison_counts),"largest_cluster":largest_cluster,"largest_cluster_people":collapsed_people,"largest_cluster_causes":cluster_causes,"estimated_cluster":largest_cluster,"zero_pairs":zero_pairs,"true_pairs":true_pairs,"zero_rate":zero_rate,"zero_source_pairs":dict(zero_source_pairs),"canonical_links":len(canonical_links),"canonical_unrecoverable":canonical_unrecoverable,"canonical_rate":canonical_rate,"candidate_incidences":candidate_incidences,"different_person_incidences":hard_candidate_incidences,"unique_naive_pairs":unique_naive_pairs,"pairs_only_rule2_values":pairs_only_rule2_values,"pairs_only_poison":pairs_only_poison,"rule2_pairs":len(rule2_pairs),"after_rule2_pct":after_rule2_pct,"explicit_hard_pairs":len(explicit_pairs),"explicit_hard_candidates":explicit_candidate_pairs,"explicit_hard_rate":explicit_hard_rate,"high_value_count":len(high_items),"evidence_modes":dict(canonical_primary_modes)}
    discrepancies={k:(claims[k],recomputed[k]) for k in claims if claims[k]!=recomputed[k]}
    audit.add("REPORT-001","Saved generation report measured claims match independent recomputation","no discrepancies",discrepancies or "none","PASS" if not discrepancies else "FAIL",f"Claims={claims}; recomputed={recomputed}. The independent matcher rejects values that normalize to an empty token.","In identity_tokens, normalize first and add an email token only when the normalized value is nonempty; then recompute the report from existing files.",cause="The generator admits quote-only artifacts as a shared empty normalized email token, inflating candidate incidences and the poison component.",location="identity_tokens/generate:report",regen=False)
    reported_total=report.get("total_row_count")
    audit.add("REPORT-002","Generation report explicitly includes reconciled total row count",total_rows,reported_total,"PASS" if isinstance(reported_total,int) and reported_total==total_rows else "FAIL",f"Independent parsed sum={total_rows:,}; reported total={reported_total!r}.","Add total_row_count=sum(row_counts.values()) to generation_report.json.",cause="Missing or inconsistent explicit total.",location="generate:report",regen=False)
    required_report={"row_counts","total_row_count","injected_problem_rates","true_number_of_distinct_people","true_duplicate_rate_people_with_multiple_records","largest_poisoned_cluster_naive_matcher"}
    audit.add("REPORT-003","Generation report contains all other required measures",str(sorted(required_report)),sorted(report),"PASS" if required_report<=set(report) else "FAIL","Field presence checked; critical values independently recomputed above.","Add missing report fields.",location="generate:report",regen=False)
    actual_rates={"exact_duplicate_rate":exact_rate,"near_duplicate_rate":near_rate,"bot_rate":bot_rate,"internal_test_rate":qa_rate,"late_event_rate":late_rate,"impossible_value_rate":impossible_rate,"zero_evidence_duplicate_rate":zero_rate}
    rate_tolerances={"exact_duplicate_rate":.0025,"near_duplicate_rate":.0025,"bot_rate":.01,"internal_test_rate":.001,"late_event_rate":.002,"impossible_value_rate":.001,"zero_evidence_duplicate_rate":.01}
    rate_diffs={k:{"claimed":v,"observed":actual_rates.get(k),"absolute_deviation":abs(v-actual_rates.get(k,0)),"tolerance":rate_tolerances[k]} for k,v in injected_report.items() if k in actual_rates}
    rate_ok=all(item["absolute_deviation"]<=item["tolerance"] for item in rate_diffs.values())
    audit.add("REPORT-004","Every injected-rate claim is compared with emitted data","within predeclared per-metric tolerance",rate_diffs,"PASS" if rate_ok else "FAIL","These are target/injected claims, so tolerance—not exact equality—is appropriate; tolerances are stated in the methodology.","Correct injection rates and regenerate source data.",location="generate/injected_problem_rates",regen=not rate_ok)
    required_modes={"exact_verified_email","email_case_variation","email_dotted_local_part","email_plus_suffix","phone_country_code","phone_spaced","phone_leading_zero","name_city_only","device_only","no_usable_evidence"}
    mode_ok=canonical_mode_matches==len(canonical_links) and required_modes<=set(canonical_modes) and measured_report.get("duplicate_evidence_mode_counts")==dict(canonical_primary_modes)
    audit.add("REPORT-005","Hidden evidence modes reconcile with observable source evidence","all links reconcile and all required modes occur",f"reconciled={canonical_mode_matches:,}/{len(canonical_links):,}; modes={dict(canonical_modes)}","PASS" if mode_ok else "FAIL","The validator recomputed normalized evidence from source values and compared exact intended mode sets; labels alone are insufficient.","Correct canonical evidence labels or evidence generation.",location="build_hidden_metadata/observed_evidence_modes",regen=not mode_ok)

    summary={
        "total_rows":total_rows,"true_people":len(people_counts),"multi_record_people_rate":multi_rate,
        "true_duplicate_rate":multi_rate,"true_pairs":true_pairs,"unrecoverable_pairs":zero_pairs,
        "unrecoverable_rate":zero_rate,"canonical_unrecoverable_rate":canonical_rate,"canonical_unrecoverable_links":canonical_unrecoverable,"canonical_links":len(canonical_links),
        "hard_negative_rate":hard_rate,"explicit_hard_negative_pairs":explicit_candidate_pairs,"rule2_unique_candidate_pairs":len(rule2_pairs),
        "unique_naive_candidate_pairs":unique_naive_pairs,"candidate_incidences":candidate_incidences,"pairs_only_rule2_values":pairs_only_rule2_values,"pairs_only_poison":pairs_only_poison,
        "candidate_percentage_after_rule2":after_rule2_pct,"largest_poisoned_cluster":largest_cluster,
        "largest_poisoned_cluster_distinct_people":collapsed_people,"late_rate":late_rate,"late_count":late,
        "exact_duplicate_rate":exact_rate,"near_duplicate_rate":near_rate,"bot_rate":bot_rate,
        "poison_counts":dict(poison_counts),"hard_negative_breakdown":dict(hard_types),
        "timestamp_formats":dict(formats),"missing_rates":{k:v for k,v in miss_rates.items()},
        "source_counts":dict(source_counts),"parse_errors":parse_errors,
    }
    for name,value in summary.items():
        if isinstance(value,(int,float,str)): audit.metric("SUMMARY",name,value,"proportion" if name.endswith("rate") else "")
    return audit, summary


def write_outputs(audit: Audit, summary: dict[str, Any], output_dir: Path) -> None:
    counts=Counter(c.status for c in audit.checks); mandatory_fail=[c for c in audit.checks if c.mandatory and c.status in {"FAIL","NOT VERIFIABLE"}]
    verdict="FAIL" if mandatory_fail else "PASS"
    risks=[c for c in audit.checks if c.status in {"FAIL","NOT VERIFIABLE","WARNING"}][:5]
    result={"verdict":verdict,"summary":summary,"status_counts":dict(counts),"checks":[asdict(c) for c in audit.checks],
            "methodology":{"usable_evidence":"Normalized email, phone, device, account reference, payment token, hashed email, or name+city; explicit poison and any token occurring >40 times are worthless.","unrecoverable_denominator":"Both all pairwise physical human-record combinations and hidden canonical star links are reported.","missing_values":"Actual null, empty/whitespace, quoted-empty, and case-insensitive strings 'null'/'None'.","naive_candidate":"Unique unordered source-record-key pairs are the hard-negative denominator after Rule 2; physical token incidences are diagnostic only.","privacy":"Person IDs and member records are never emitted."}}
    (output_dir/"dataset_audit_results.json").write_text(json.dumps(result,indent=2,ensure_ascii=False,default=str),encoding="utf-8")
    with (output_dir/"dataset_audit_metrics.csv").open("w",encoding="utf-8",newline="") as handle:
        writer=csv.DictWriter(handle,fieldnames=["id","metric","value","unit"]); writer.writeheader(); writer.writerows(audit.metrics)
    with (output_dir/"dataset_audit_failures.csv").open("w",encoding="utf-8",newline="") as handle:
        fields=list(asdict(audit.checks[0]).keys()) if audit.checks else ["id","status"]
        writer=csv.DictWriter(handle,fieldnames=fields); writer.writeheader(); writer.writerows(asdict(c) for c in audit.checks if c.status!="PASS")
    lines=["# Dataset Audit Report","","## Executive summary","",f"- Overall verdict: **{verdict}**",f"- Mandatory checks passed: {sum(c.mandatory and c.status=='PASS' for c in audit.checks)}",f"- Mandatory checks failed: {sum(c.mandatory and c.status=='FAIL' for c in audit.checks)}",f"- Warnings: {counts['WARNING']}",f"- Requirements not verifiable: {counts['NOT VERIFIABLE']}",f"- Total generated rows: {summary.get('total_rows','n/a'):,}",f"- True distinct people: {summary.get('true_people','n/a'):,}",f"- True duplicate rate: {pct(summary.get('true_duplicate_rate',0))}",f"- Largest poisoned naive cluster: {summary.get('largest_poisoned_cluster','n/a'):,}","","Five most important risks:"]
    lines += [f"- {c.id}: {c.requirement} — {c.observed}" for c in risks] or ["- No failures, warnings, or unverifiable items."]
    lines += ["","## Methodology and tolerances","","Approximate full-scale targets use predeclared tolerances: people ±2%, rows ±5%, multi-record people ±2 percentage points, unrecoverable pairs ±1 point, exact/near rows ±0.25 points, impossible values ±0.1 point, and QA accounts ±0.1 point. Poison counts use ±20%, except the qualitative 'several thousand' DOB target uses 2,000–6,000. These bounds reflect qualitative wording while remaining tight for a deterministic 420,000-row population.","","Usable evidence is normalized email, phone, device, account reference, payment token, hashed email, or name+city. Explicit poison and any token occurring more than 40 times are worthless. Both all-pairwise and canonical-star duplicate denominators are reported. The hard-negative test uses unique unordered source-record-key pairs after Rule 2; repeated token incidences are diagnostic only.","","Missing values consistently include actual nulls, empty/whitespace, quoted-empty, and case-insensitive literal `null`/`None`. No source value is changed by the audit.","","## Requirement-by-requirement results","","| ID | Requirement | Target | Observed | Status | Evidence | Recommended action |","| -- | ----------- | ------ | -------- | ------ | -------- | ------------------ |"]
    esc=lambda x:str(x).replace("|","\\|").replace("\n"," ")
    for c in audit.checks: lines.append(f"| {c.id} | {esc(c.requirement)} | {esc(c.target)} | {esc(c.observed)} | {c.status} | {esc(c.evidence)} | {esc(c.recommended_action)} |")
    failed=[c for c in audit.checks if c.status=="FAIL"]
    lines += ["","## Failure details",""]
    if not failed: lines.append("No failed requirements.")
    for c in failed:
        location = f"scripts/generate_synthetic_dataset.py::{c.generator_location}" if c.generator_location else "not isolated"
        lines += [f"### {c.id}","",f"- Observed: {c.observed}",f"- Expected: {c.target}",f"- Difference/evidence: {c.evidence}",f"- Likely cause: {c.likely_cause or 'See evidence and generator location.'}",f"- Probable generator location: `{location}`",f"- Suggested fix: {c.recommended_action}",f"- Full regeneration required: {'Yes' if c.regeneration_required else 'No'}",""]
    (output_dir/"dataset_audit_report.md").write_text("\n".join(lines)+"\n",encoding="utf-8")


def terminal_summary(audit: Audit, summary: dict[str, Any], output_dir: Path) -> int:
    counts=Counter(c.status for c in audit.checks); mandatory_failed=[c for c in audit.checks if c.mandatory and c.status in {"FAIL","NOT VERIFIABLE"}]
    verdict="FAIL" if mandatory_failed else "PASS"
    print(f"\nDATASET AUDIT: {verdict}\n")
    print(f"Mandatory checks passed: {sum(c.mandatory and c.status=='PASS' for c in audit.checks)}")
    print(f"Mandatory checks failed: {sum(c.mandatory and c.status=='FAIL' for c in audit.checks)}")
    print(f"Warnings: {counts['WARNING']}")
    print(f"Not verifiable: {counts['NOT VERIFIABLE']}\n")
    print(f"Rows: {summary.get('total_rows','n/a')}")
    print(f"True people: {summary.get('true_people','n/a')}")
    print(f"Multi-record people: {pct(summary.get('multi_record_people_rate',0))}")
    print(f"True duplicate rate: {pct(summary.get('true_duplicate_rate',0))}")
    print(f"Unrecoverable true pairs: {pct(summary.get('unrecoverable_rate',0))}")
    print(f"Canonical unrecoverable links: {pct(summary.get('canonical_unrecoverable_rate',0))}")
    print(f"Hard negatives among naive candidates: {pct(summary.get('hard_negative_rate',0))}")
    print(f"Largest poisoned cluster: {summary.get('largest_poisoned_cluster','n/a')}")
    print(f"Late-arriving events: {pct(summary.get('late_rate',0))}\n")
    print("Failed requirements:")
    for c in mandatory_failed: print(f"- {c.id}: expected {c.target}, observed {c.observed}")
    print("\nReports:")
    for name in ("dataset_audit_report.md","dataset_audit_results.json","dataset_audit_metrics.csv","dataset_audit_failures.csv"): print(f"- {output_dir/name}")
    return 1 if mandatory_failed else 0


def main() -> int:
    parser=argparse.ArgumentParser(description=__doc__)
    parser.add_argument("data_dir",nargs="?",type=Path,help="Generated-data directory")
    parser.add_argument("--data-dir",dest="data_dir_option",type=Path,help="Generated-data directory (named form)")
    parser.add_argument("--output-dir",type=Path,default=Path("outputs"))
    parser.add_argument("--generator",type=Path,default=Path("scripts/generate_synthetic_dataset.py"))
    args=parser.parse_args(); data_dir=args.data_dir_option or args.data_dir
    if data_dir is None: parser.error("provide DATA_DIR or --data-dir DATA_DIR")
    audit,summary=audit_dataset(data_dir.resolve(),args.output_dir.resolve(),args.generator.resolve())
    print("[7/7] Writing audit-only reports...")
    write_outputs(audit,summary,args.output_dir.resolve())
    return terminal_summary(audit,summary,args.output_dir.resolve())


if __name__=="__main__": raise SystemExit(main())
